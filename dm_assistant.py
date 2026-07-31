from __future__ import annotations

import json
import mimetypes
import os
import re
import shutil
import subprocess
import threading
import time
import urllib.error
import urllib.request
from copy import copy
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from email.policy import SMTP
from email.utils import make_msgid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse
from uuid import uuid4

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from calendar_sync import fetch_sheet_events, merge_sheet_events
from brand_connect_crawler import BrandConnectCrawlerManager, normalize_creator
from brand_connect_favorite import (
    PLATFORM_LABELS as FAVORITE_PLATFORM_LABELS,
    PRODUCT_LABELS as FAVORITE_PRODUCT_LABELS,
    RUNNING_STATUSES as FAVORITE_RUNNING_STATUSES,
    BrandConnectFavoriteManager,
)
from brand_connect_proposal import (
    PRODUCT_LABELS as PROPOSAL_PRODUCT_LABELS,
    RUNNING_STATUSES as PROPOSAL_RUNNING_STATUSES,
    BrandConnectProposalManager,
    normalize_proposal_date,
)
from brand_connect_sheet import (
    PROPOSAL_STATUS_VALUES,
    find_favorite_date_column,
    is_favorite_candidate,
)
from dm_templates import DM_MESSAGE_TEMPLATES
from instagram_dm_sender import InstagramDMSenderManager
from price_updater import PriceUpdateManager
from simulation_manager import SimulationManager


APP_DIR = Path(__file__).resolve().parent
CONFIG_PATH = APP_DIR / "config.json"
HTML_PATH = APP_DIR / "index.html"
PRICE_HTML_PATH = APP_DIR / "price.html"
SIMULATION_HTML_PATH = APP_DIR / "simulation.html"
CALENDAR_HTML_PATH = APP_DIR / "calendar.html"
MEETINGS_HTML_PATH = APP_DIR / "meetings.html"
MEMOS_HTML_PATH = APP_DIR / "memos.html"
SALES_EMAIL_HTML_PATH = APP_DIR / "sales_email.html"
BRAND_CONNECTING_HTML_PATH = APP_DIR / "brand_connecting.html"
THEME_CSS_PATH = APP_DIR / "theme_meeting.css"
BRAND_CONNECT_CAMPAIGNS_PATH = (
    APP_DIR / "app_data" / "brand_connect_campaigns.json"
)
BRAND_CONNECTING_SHEETS = {
    "alp": {
        "brand_name": "알프",
        "sheet_name": "알프 쇼핑커넥트",
        "column_count": 9,
    },
    "gaia": {
        "brand_name": "가이아",
        "sheet_name": "가이아 쇼핑커넥트",
        "column_count": 12,
    },
}
OUTLOOK_POPUP_SCRIPT_PATH = APP_DIR / "open_new_outlook_draft.ps1"
SKIP_LOG_PATH = APP_DIR / "skip_log.json"
CALENDAR_TASKS_PATH = APP_DIR / "calendar_tasks.json"
MEETING_NOTES_PATH = APP_DIR / "meeting_notes.json"
PLATFORM_MEMOS_PATH = APP_DIR / "platform_memos.json"
SALES_EMAIL_DATA_PATH = APP_DIR / "sales_email_data.json"
SALES_EMAIL_ATTACHMENTS_DIR = APP_DIR / "sales_email_attachments"
DM_REFERENCE_IMAGE_DIR = APP_DIR / "sales_email_assets"
DM_REFERENCE_IMAGES = {
    "gaia": DM_REFERENCE_IMAGE_DIR / "가이아_공동구매_구성제안.png",
    "alp": DM_REFERENCE_IMAGE_DIR / "알프_공동구매_구성안.png",
}
DM_TEMPLATE_OVERRIDES_PATH = APP_DIR / "app_data" / "dm_template_overrides.json"
MEETING_RECORDINGS_DIR = APP_DIR / "meeting_recordings"
LOCAL_WHISPER_DIR = Path(
    os.environ.get(
        "TOOBUS_WHISPER_DIR",
        str(APP_DIR / "local_models" / "whisper"),
    )
)
LOCAL_WHISPER_MODEL = "small"
LOCAL_SUMMARY_MODEL = "qwen2.5:7b"
HOST = "127.0.0.1"
PORT = 8765
MAX_RECORDING_BYTES = 1024 * 1024 * 1024


def load_config() -> dict:
    with CONFIG_PATH.open("r", encoding="utf-8") as file:
        config = json.load(file)
    for brand_key, message_template in DM_MESSAGE_TEMPLATES.items():
        brand = config.get("brands", {}).get(brand_key)
        if isinstance(brand, dict):
            brand["message_template"] = message_template
    workbook_path = Path(config["workbook_path"])
    if not workbook_path.is_absolute():
        workbook_path = APP_DIR / workbook_path
    config["_workbook_path"] = workbook_path.resolve()
    return config


CONFIG = load_config()
PRICE_MANAGER = PriceUpdateManager(APP_DIR, CONFIG["price_updater"])
SIMULATION_MANAGER = SimulationManager(
    APP_DIR, CONFIG["simulation"], PRICE_MANAGER
)
LOCK = threading.Lock()
DM_SYNC_LOCK = threading.Lock()
DM_TEMPLATE_LOCK = threading.Lock()
CALENDAR_LOCK = threading.Lock()
CALENDAR_SYNC_LOCK = threading.Lock()
MEETING_NOTES_LOCK = threading.Lock()
PLATFORM_MEMOS_LOCK = threading.Lock()
SALES_EMAIL_LOCK = threading.Lock()
MEETING_AI_LOCK = threading.Lock()
MEETING_AI_JOBS_LOCK = threading.Lock()
MEETING_AI_JOBS: dict[str, dict] = {}
WHISPER_MODEL_INSTANCE = None
BACKUP_CREATED = False
DM_SYNC_STATE_PATH = APP_DIR / "app_data" / "dm_sync_state.json"
DM_DRIVE_TOKEN_PATH = APP_DIR / "app_data" / "dm_drive_token.json"
DM_SYNC_STATUS = {
    "status": "idle",
    "message": "온라인 원본 확인 전",
    "last_checked_at": None,
    "last_synced_at": None,
    "last_modified": None,
}
CALENDAR_SYNC_STATUS = {
    "status": "idle",
    "message": "아직 동기화하지 않았습니다.",
    "last_synced_at": None,
}
SESSION = {
    brand_key: {"sent": 0, "skipped": 0}
    for brand_key in CONFIG["brands"]
}


def load_calendar_tasks() -> list[dict]:
    if not CALENDAR_TASKS_PATH.exists():
        return []
    try:
        data = json.loads(CALENDAR_TASKS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    return data if isinstance(data, list) else []


def save_calendar_tasks(tasks: list[dict]) -> None:
    temporary = CALENDAR_TASKS_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(tasks, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(CALENDAR_TASKS_PATH)


def load_meeting_notes() -> list[dict]:
    if not MEETING_NOTES_PATH.exists():
        return []
    try:
        data = json.loads(MEETING_NOTES_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    return data if isinstance(data, list) else []


def save_meeting_notes(notes: list[dict]) -> None:
    temporary = MEETING_NOTES_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(notes, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(MEETING_NOTES_PATH)


def load_platform_memos() -> list[dict]:
    if not PLATFORM_MEMOS_PATH.exists():
        return []
    try:
        data = json.loads(PLATFORM_MEMOS_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return []
    return data if isinstance(data, list) else []


def save_platform_memos(memos: list[dict]) -> None:
    temporary = PLATFORM_MEMOS_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(memos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(PLATFORM_MEMOS_PATH)


def default_sales_email_data() -> dict:
    return {
        "contacts": [],
        "templates": [],
        "drafts": {},
        "history": [],
    }


def load_sales_email_data() -> dict:
    if not SALES_EMAIL_DATA_PATH.exists():
        return default_sales_email_data()
    try:
        data = json.loads(SALES_EMAIL_DATA_PATH.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return default_sales_email_data()
    if not isinstance(data, dict):
        return default_sales_email_data()
    defaults = default_sales_email_data()
    for key, value in defaults.items():
        if not isinstance(data.get(key), type(value)):
            data[key] = value
    return data


def save_sales_email_data(data: dict) -> None:
    temporary = SALES_EMAIL_DATA_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    temporary.replace(SALES_EMAIL_DATA_PATH)


def normalize_sales_email_cc(value) -> list[str]:
    if isinstance(value, str):
        candidates = re.split(r"[,;\n]+", value)
    elif isinstance(value, (list, tuple)):
        candidates = value
    else:
        candidates = []
    addresses = []
    seen = set()
    for item in candidates:
        address = str(item).strip()
        if not address:
            continue
        if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", address):
            raise ValueError(f"추가 참조 이메일 형식이 올바르지 않습니다: {address}")
        key = address.lower()
        if key in seen:
            continue
        seen.add(key)
        addresses.append(address[:320])
        if len(addresses) >= 20:
            break
    return addresses


def public_sales_email_data() -> dict:
    with SALES_EMAIL_LOCK:
        data = load_sales_email_data()
    data["history"] = sorted(
        data["history"],
        key=lambda item: str(item.get("sent_at", "")),
        reverse=True,
    )[:200]
    return data


def update_sales_email_data(payload: dict) -> dict:
    action = str(payload.get("action", "")).strip()
    now = datetime.now().isoformat(timespec="seconds")
    with SALES_EMAIL_LOCK:
        data = load_sales_email_data()
        if action == "save_contact":
            contact_id = str(payload.get("id", "")).strip()
            email = str(payload.get("email", "")).strip()
            name = str(payload.get("name", "")).strip()
            company = str(payload.get("company", "")).strip()
            platform = str(payload.get("platform", "")).strip()
            group = str(payload.get("group", "")).strip()
            memo = str(payload.get("memo", "")).strip()
            if not name:
                raise ValueError("담당자명을 입력해주세요.")
            if not re.fullmatch(r"[^@\s]+@[^@\s]+\.[^@\s]+", email):
                raise ValueError("올바른 이메일 주소를 입력해주세요.")
            contact = next(
                (item for item in data["contacts"] if item.get("id") == contact_id),
                None,
            )
            if contact is None:
                contact = {"id": uuid4().hex, "created_at": now}
                data["contacts"].append(contact)
            contact.update(
                {
                    "name": name[:100],
                    "email": email[:320],
                    "company": company[:150],
                    "platform": platform[:100],
                    "group": group[:100],
                    "memo": memo[:3000],
                    "updated_at": now,
                }
            )
            result = {"contact": dict(contact)}
        elif action == "delete_contact":
            contact_id = str(payload.get("id", "")).strip()
            if not any(item.get("id") == contact_id for item in data["contacts"]):
                raise ValueError("삭제할 연락처를 찾을 수 없습니다.")
            data["contacts"] = [
                item for item in data["contacts"] if item.get("id") != contact_id
            ]
            data["drafts"].pop(contact_id, None)
            result = {"id": contact_id}
        elif action == "save_template":
            template_id = str(payload.get("id", "")).strip()
            name = str(payload.get("name", "")).strip()
            subject = str(payload.get("subject", "")).strip()
            body = str(payload.get("body", "")).strip()
            extra_cc = normalize_sales_email_cc(payload.get("extra_cc", []))
            if not name or not subject or not body:
                raise ValueError("양식명, 제목, 본문을 모두 입력해주세요.")
            template = next(
                (item for item in data["templates"] if item.get("id") == template_id),
                None,
            )
            if template is None:
                template = {"id": uuid4().hex, "created_at": now}
                data["templates"].append(template)
            template.update(
                {
                    "name": name[:120],
                    "subject": subject[:500],
                    "body": body[:30000],
                    "extra_cc": extra_cc,
                    "updated_at": now,
                }
            )
            result = {"template": dict(template)}
        elif action == "delete_template":
            template_id = str(payload.get("id", "")).strip()
            data["templates"] = [
                item for item in data["templates"] if item.get("id") != template_id
            ]
            result = {"id": template_id}
        elif action == "save_draft":
            contact_id = str(payload.get("contact_id", "")).strip()
            if not any(item.get("id") == contact_id for item in data["contacts"]):
                raise ValueError("연락처를 먼저 선택해주세요.")
            draft = {
                "contact_id": contact_id,
                "template_id": str(payload.get("template_id", "")).strip(),
                "subject": str(payload.get("subject", "")).strip()[:500],
                "body": str(payload.get("body", "")).strip()[:30000],
                "extra_cc": normalize_sales_email_cc(payload.get("extra_cc", [])),
                "attachments": payload.get("attachments", [])[:20],
                "updated_at": now,
            }
            data["drafts"][contact_id] = draft
            result = {"draft": dict(draft)}
        elif action == "record_sent":
            contact_id = str(payload.get("contact_id", "")).strip()
            contact = next(
                (item for item in data["contacts"] if item.get("id") == contact_id),
                None,
            )
            if contact is None:
                raise ValueError("연락처를 먼저 선택해주세요.")
            subject = str(payload.get("subject", "")).strip()
            body = str(payload.get("body", "")).strip()
            if not subject or not body:
                raise ValueError("메일 제목과 본문을 입력해주세요.")
            record = {
                "id": uuid4().hex,
                "contact_id": contact_id,
                "contact_name": contact.get("name", ""),
                "email": contact.get("email", ""),
                "cc": normalize_sales_email_cc(payload.get("cc", [])),
                "extra_cc": normalize_sales_email_cc(payload.get("extra_cc", [])),
                "template_id": str(payload.get("template_id", "")).strip(),
                "subject": subject[:500],
                "body": body[:30000],
                "full_body": str(payload.get("full_body", body)).strip()[:40000],
                "full_html": str(payload.get("full_html", "")).strip()[:60000],
                "attachments": payload.get("attachments", [])[:20],
                "sent_at": now,
            }
            data["history"].append(record)
            data["drafts"][contact_id] = {
                "contact_id": contact_id,
                "template_id": record["template_id"],
                "subject": record["subject"],
                "body": record["body"],
                "extra_cc": record["extra_cc"],
                "attachments": record["attachments"],
                "updated_at": now,
                "last_sent_at": now,
            }
            result = {"record": record}
        else:
            raise ValueError("지원하지 않는 영업메일 작업입니다.")
        save_sales_email_data(data)
    return {"ok": True, **result}


def save_sales_email_attachment(
    filename: str,
    mime_type: str,
    source,
    length: int,
) -> dict:
    if length <= 0:
        raise ValueError("첨부할 파일을 선택해주세요.")
    if length > 25 * 1024 * 1024:
        raise ValueError("첨부파일은 한 개당 25MB까지 가능합니다.")
    safe_name = Path(filename).name.strip() or "attachment"
    attachment_id = uuid4().hex
    directory = SALES_EMAIL_ATTACHMENTS_DIR / attachment_id
    directory.mkdir(parents=True, exist_ok=True)
    saved_path = directory / safe_name
    remaining = length
    with saved_path.open("wb") as output:
        while remaining:
            chunk = source.read(min(1024 * 1024, remaining))
            if not chunk:
                break
            output.write(chunk)
            remaining -= len(chunk)
    if remaining:
        shutil.rmtree(directory, ignore_errors=True)
        raise ValueError("첨부파일 저장이 중단되었습니다.")
    return {
        "ok": True,
        "attachment": {
            "id": attachment_id,
            "name": safe_name,
            "size": length,
            "mime_type": mime_type or "application/octet-stream",
        },
    }


def create_outlook_draft(payload: dict) -> dict:
    contact_id = str(payload.get("contact_id", "")).strip()
    with SALES_EMAIL_LOCK:
        data = load_sales_email_data()
        contact = next(
            (item for item in data["contacts"] if item.get("id") == contact_id),
            None,
        )
    if contact is None:
        raise ValueError("연락처를 먼저 선택해주세요.")
    subject = str(payload.get("subject", "")).strip()
    plain_body = str(payload.get("full_body", "")).strip()
    html_body = str(payload.get("full_html", "")).strip()
    recipient = str(contact.get("email", "")).strip()
    if not subject or not plain_body or not html_body or not recipient:
        raise ValueError("메일 제목과 본문을 입력해주세요.")
    cc_addresses = normalize_sales_email_cc(payload.get("cc", []))
    attachment_paths = []
    attachment_root = SALES_EMAIL_ATTACHMENTS_DIR.resolve()
    for attachment in payload.get("attachments", [])[:20]:
        attachment_id = str(attachment.get("id", "")).strip()
        name = Path(str(attachment.get("name", ""))).name
        candidate = (SALES_EMAIL_ATTACHMENTS_DIR / attachment_id / name).resolve()
        if candidate.is_file() and candidate.is_relative_to(attachment_root):
            attachment_paths.append(str(candidate))
    temporary_dir = APP_DIR / "app_data" / "outlook_drafts"
    temporary_dir.mkdir(parents=True, exist_ok=True)
    message = EmailMessage(policy=SMTP)
    message["X-Unsent"] = "1"
    message["Message-ID"] = make_msgid(domain="tobuscorp.com")
    message["To"] = recipient
    if cc_addresses:
        message["Cc"] = ", ".join(cc_addresses)
    message["Subject"] = subject[:500]
    message.set_content(plain_body[:30000], charset="utf-8")
    message.add_alternative(html_body[:60000], subtype="html", charset="utf-8")
    for attachment_path in attachment_paths:
        file_path = Path(attachment_path)
        mime_type, _ = mimetypes.guess_type(file_path.name)
        main_type, sub_type = (
            mime_type.split("/", 1)
            if mime_type and "/" in mime_type
            else ("application", "octet-stream")
        )
        message.add_attachment(
            file_path.read_bytes(),
            maintype=main_type,
            subtype=sub_type,
            filename=file_path.name,
        )
    eml_path = temporary_dir / f"TOBUS_{uuid4().hex}.eml"
    eml_path.write_bytes(message.as_bytes())
    if not OUTLOOK_POPUP_SCRIPT_PATH.is_file():
        raise ValueError("Outlook 팝업 실행 파일을 찾을 수 없습니다.")
    creation_flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    completed = subprocess.run(
        [
            "C:\\WINDOWS\\System32\\WindowsPowerShell\\v1.0\\powershell.exe",
            "-NoProfile",
            "-STA",
            "-ExecutionPolicy",
            "Bypass",
            "-WindowStyle",
            "Hidden",
            "-File",
            str(OUTLOOK_POPUP_SCRIPT_PATH),
            "-DraftPath",
            str(eml_path),
        ],
        capture_output=True,
        text=True,
        timeout=15,
        creationflags=creation_flags,
    )
    if completed.returncode != 0:
        detail = (completed.stderr or completed.stdout or "").strip()
        raise ValueError(
            "새 Outlook 팝업을 열지 못했습니다."
            + (f" {detail[-400:]}" if detail else "")
        )
    return {
        "ok": True,
        "message": "본문·첨부파일·HTML 로고 서명이 포함된 Outlook 초안을 열었습니다.",
        "attachments": len(attachment_paths),
    }


def get_platform_memos() -> dict:
    with PLATFORM_MEMOS_LOCK:
        memos = load_platform_memos()
    memos.sort(
        key=lambda item: str(item.get("updated_at", item.get("created_at", ""))),
        reverse=True,
    )
    return {"memos": memos}


def update_platform_memo(payload: dict) -> dict:
    action = str(payload.get("action", "")).strip()
    with PLATFORM_MEMOS_LOCK:
        memos = load_platform_memos()
        if action == "save":
            memo_id = str(payload.get("id", "")).strip()
            platform = str(payload.get("platform", "")).strip()
            content = str(payload.get("memo", "")).strip()
            if not platform:
                raise ValueError("플랫폼명을 입력해주세요.")
            if not content:
                raise ValueError("메모 내용을 입력해주세요.")
            if len(platform) > 100:
                raise ValueError("플랫폼명은 100자까지 입력할 수 있습니다.")
            if len(content) > 5000:
                raise ValueError("메모는 5,000자까지 입력할 수 있습니다.")
            now = datetime.now().isoformat(timespec="seconds")
            memo = next(
                (item for item in memos if item.get("id") == memo_id),
                None,
            )
            if memo is None:
                memo = {
                    "id": uuid4().hex,
                    "created_at": now,
                }
                memos.append(memo)
            memo.update(
                {
                    "platform": platform,
                    "memo": content,
                    "updated_at": now,
                }
            )
            response_memo = dict(memo)
        elif action == "delete":
            memo_id = str(payload.get("id", "")).strip()
            if not any(item.get("id") == memo_id for item in memos):
                raise ValueError("삭제할 메모를 찾을 수 없습니다.")
            memos = [item for item in memos if item.get("id") != memo_id]
            response_memo = {"id": memo_id}
        else:
            raise ValueError("지원하지 않는 메모 작업입니다.")
        save_platform_memos(memos)
    return {"ok": True, "memo": response_memo}


def ollama_executable() -> Path | None:
    candidates = [
        Path(os.environ.get("LOCALAPPDATA", ""))
        / "Programs"
        / "Ollama"
        / "ollama.exe",
        Path("C:/Program Files/Ollama/ollama.exe"),
    ]
    return next((path for path in candidates if path.is_file()), None)


def ollama_models() -> list[str]:
    try:
        with urllib.request.urlopen(
            "http://127.0.0.1:11434/api/tags",
            timeout=3,
        ) as response:
            payload = json.loads(response.read().decode("utf-8"))
        return [str(item.get("name", "")) for item in payload.get("models", [])]
    except (OSError, urllib.error.URLError, json.JSONDecodeError):
        return []


def ensure_ollama_running() -> None:
    if ollama_models():
        return
    executable = ollama_executable()
    if executable is None:
        raise ValueError("무료 회의록 정리 모델 실행기가 설치되지 않았습니다.")
    creation_flags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    subprocess.Popen(
        [str(executable), "serve"],
        cwd=str(executable.parent),
        stdout=subprocess.DEVNULL,
        stderr=subprocess.DEVNULL,
        creationflags=creation_flags,
    )
    for _ in range(20):
        time.sleep(0.25)
        if ollama_models():
            return
    raise ValueError("무료 회의록 정리 모델을 시작하지 못했습니다.")


def meeting_ai_status() -> dict:
    try:
        import faster_whisper  # noqa: F401

        whisper_installed = True
    except ImportError:
        whisper_installed = False
    whisper_ready = LOCAL_WHISPER_DIR.exists() and any(
        LOCAL_WHISPER_DIR.rglob("model.bin")
    )
    models = ollama_models()
    if not models and ollama_executable() is not None:
        try:
            ensure_ollama_running()
        except ValueError:
            pass
        models = ollama_models()
    summary_ready = any(
        name == LOCAL_SUMMARY_MODEL or name.startswith(f"{LOCAL_SUMMARY_MODEL}:")
        for name in models
    )
    return {
        "configured": whisper_installed and whisper_ready and summary_ready,
        "free": True,
        "offline": True,
        "whisper_ready": whisper_ready,
        "summary_ready": summary_ready,
        "transcription_model": f"faster-whisper {LOCAL_WHISPER_MODEL}",
        "summary_model": LOCAL_SUMMARY_MODEL,
    }


def recording_extension(mime_type: str) -> str:
    normalized = mime_type.split(";", 1)[0].strip().lower()
    return {
        "audio/webm": ".webm",
        "audio/ogg": ".ogg",
        "audio/mp4": ".m4a",
        "video/mp4": ".mp4",
        "audio/mpeg": ".mp3",
        "audio/wav": ".wav",
        "audio/x-wav": ".wav",
    }.get(normalized, "")


def find_meeting_note(notes: list[dict], note_id: str) -> dict:
    note = next((item for item in notes if item.get("id") == note_id), None)
    if note is None:
        raise ValueError("회의록을 먼저 저장해주세요.")
    return note


def save_meeting_recording(
    note_id: str,
    mime_type: str,
    duration_seconds: int,
    stream,
    content_length: int,
    source: str = "meeting",
) -> dict:
    extension = recording_extension(mime_type)
    if not extension:
        raise ValueError("지원하지 않는 녹음 형식입니다.")
    if content_length <= 0:
        raise ValueError("녹음 내용이 비어 있습니다.")
    if content_length > MAX_RECORDING_BYTES:
        raise ValueError("녹음 파일이 너무 큽니다.")
    recording_id = uuid4().hex
    directory = MEETING_RECORDINGS_DIR / note_id
    directory.mkdir(parents=True, exist_ok=True)
    destination = directory / f"{recording_id}{extension}"
    temporary = destination.with_suffix(f"{extension}.tmp")
    remaining = content_length
    with temporary.open("wb") as output:
        while remaining:
            chunk = stream.read(min(1024 * 1024, remaining))
            if not chunk:
                raise ValueError("녹음 파일 전송이 중간에 끊겼습니다.")
            output.write(chunk)
            remaining -= len(chunk)
    temporary.replace(destination)
    recording = {
        "id": recording_id,
        "filename": destination.name,
        "mime_type": mime_type.split(";", 1)[0].strip().lower(),
        "size": content_length,
        "duration_seconds": max(0, int(duration_seconds)),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "transcript": "",
        "source": source if source in {"meeting", "teams", "import"} else "meeting",
    }
    try:
        with MEETING_NOTES_LOCK:
            notes = load_meeting_notes()
            note = find_meeting_note(notes, note_id)
            note.setdefault("recordings", []).append(recording)
            note["updated_at"] = datetime.now().isoformat(timespec="seconds")
            save_meeting_notes(notes)
    except Exception:
        destination.unlink(missing_ok=True)
        raise
    return {"ok": True, "recording": recording}


def delete_meeting_recording(note_id: str, recording_id: str) -> dict:
    with MEETING_NOTES_LOCK:
        notes = load_meeting_notes()
        note = find_meeting_note(notes, note_id)
        recordings = note.get("recordings", [])
        recording = next(
            (item for item in recordings if item.get("id") == recording_id),
            None,
        )
        if recording is None:
            raise ValueError("삭제할 녹음을 찾을 수 없습니다.")
        note["recordings"] = [
            item for item in recordings if item.get("id") != recording_id
        ]
        note["updated_at"] = datetime.now().isoformat(timespec="seconds")
        save_meeting_notes(notes)
    path = MEETING_RECORDINGS_DIR / note_id / str(recording.get("filename", ""))
    path.unlink(missing_ok=True)
    return {"ok": True, "recording": {"id": recording_id}}


def get_whisper_model():
    global WHISPER_MODEL_INSTANCE
    if WHISPER_MODEL_INSTANCE is None:
        try:
            from faster_whisper import WhisperModel
        except ImportError as exc:
            raise ValueError("무료 받아쓰기 엔진이 설치되지 않았습니다.") from exc
        if not LOCAL_WHISPER_DIR.exists():
            raise ValueError("무료 받아쓰기 모델을 아직 내려받는 중입니다.")
        WHISPER_MODEL_INSTANCE = WhisperModel(
            LOCAL_WHISPER_MODEL,
            device="cpu",
            compute_type="int8",
            cpu_threads=max(1, min(4, (os.cpu_count() or 4) // 2)),
            num_workers=1,
            download_root=str(LOCAL_WHISPER_DIR),
            local_files_only=True,
        )
    return WHISPER_MODEL_INSTANCE


def clock_label(seconds: float) -> str:
    total = max(0, int(seconds))
    return f"{total // 3600:02d}:{(total % 3600) // 60:02d}:{total % 60:02d}"


def transcribe_recording(path: Path) -> tuple[str, list[dict]]:
    model = get_whisper_model()
    segment_iterator, info = model.transcribe(
        str(path),
        beam_size=5,
        vad_filter=True,
        vad_parameters={"min_silence_duration_ms": 500},
        condition_on_previous_text=True,
        word_timestamps=False,
    )
    lines: list[str] = []
    clean_segments: list[dict] = []
    for segment in segment_iterator:
        text = str(segment.text).strip()
        if not text:
            continue
        clean = {
            "speaker": "",
            "text": text,
            "start": round(float(segment.start), 2),
            "end": round(float(segment.end), 2),
            "language": info.language,
        }
        clean_segments.append(clean)
        lines.append(f"[{clock_label(segment.start)}] {text}")
    transcript = "\n".join(lines)
    if not transcript:
        raise ValueError("녹음에서 또렷한 음성을 찾지 못했습니다.")
    return transcript, clean_segments


def ollama_json(prompt: str) -> dict:
    ensure_ollama_running()
    schema = {
        "type": "object",
        "properties": {
            "content": {"type": "string"},
            "decisions": {"type": "string"},
            "actions": {"type": "string"},
        },
        "required": ["content", "decisions", "actions"],
        "additionalProperties": False,
    }
    payload = {
        "model": LOCAL_SUMMARY_MODEL,
        "prompt": prompt,
        "stream": False,
        "format": schema,
        "options": {
            "temperature": 0.1,
            "num_ctx": 32768,
            "num_thread": max(1, min(4, (os.cpu_count() or 4) // 2)),
        },
    }
    request = urllib.request.Request(
        "http://127.0.0.1:11434/api/generate",
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        method="POST",
        headers={"Content-Type": "application/json; charset=utf-8"},
    )
    try:
        with urllib.request.urlopen(request, timeout=1800) as response:
            result = json.loads(response.read().decode("utf-8"))
        return json.loads(str(result.get("response", "{}")))
    except (urllib.error.URLError, TimeoutError) as exc:
        raise ValueError("무료 로컬 회의록 모델이 응답하지 않습니다.") from exc
    except json.JSONDecodeError as exc:
        raise ValueError("무료 로컬 모델의 회의록 결과를 읽지 못했습니다.") from exc


def normalize_relative_time_for_summary(transcript: str) -> str:
    """Keep relative English dates relative before sending text to the local LLM."""
    weekdays = {
        "monday": "월요일",
        "tuesday": "화요일",
        "wednesday": "수요일",
        "thursday": "목요일",
        "friday": "금요일",
        "saturday": "토요일",
        "sunday": "일요일",
    }
    normalized = transcript
    replacements = (
        (r"\bby\s+tomorrow\b", "내일까지"),
        (r"\bthe\s+day\s+after\s+tomorrow\b", "모레"),
        (r"\btomorrow\b", "내일"),
        (r"\btoday\b", "오늘"),
        (r"\byesterday\b", "어제"),
        (r"\bnext\s+week\b", "다음 주"),
        (r"\bthis\s+week\b", "이번 주"),
        (r"\bnext\s+month\b", "다음 달"),
        (r"\bthis\s+month\b", "이번 달"),
    )
    for pattern, replacement in replacements:
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)
    for english, korean in weekdays.items():
        normalized = re.sub(
            rf"\bnext\s+{english}\b", f"다음 {korean}", normalized, flags=re.IGNORECASE
        )
        normalized = re.sub(
            rf"\bthis\s+{english}\b", f"이번 {korean}", normalized, flags=re.IGNORECASE
        )
        normalized = re.sub(
            rf"\bby\s+{english}\b", f"{korean}까지", normalized, flags=re.IGNORECASE
        )
    return normalized


def relative_time_evidence(transcript: str) -> list[tuple[str, str]]:
    normalized = normalize_relative_time_for_summary(transcript)
    time_pattern = re.compile(
        r"(?:다음|이번)\s+(?:월|화|수|목|금|토|일)요일|"
        r"(?:월|화|수|목|금|토|일)요일까지|"
        r"내일까지|내일|오늘|어제|모레|다음\s+(?:주|달)|이번\s+(?:주|달)"
    )
    evidence: list[tuple[str, str]] = []
    for line in normalized.splitlines():
        for match in time_pattern.finditer(line):
            evidence.append((line, match.group(0)))
    return evidence


def restore_relative_times(summary: dict, transcript: str) -> dict:
    evidence = relative_time_evidence(transcript)
    if not evidence:
        return summary
    unknown_date = "원문 날짜 확인 필요"
    action_lines = str(summary.get("actions", "")).splitlines()
    repaired_actions: list[str] = []
    for action in action_lines:
        repaired = action
        if "|" in action:
            assignee = action.lstrip("- ").split("|", 1)[0].strip()
            matching = [
                expression
                for source_line, expression in evidence
                if assignee and assignee.casefold() in source_line.casefold()
            ]
            if matching:
                fields = [field.strip() for field in action.split("|")]
                if len(fields) >= 3:
                    fields[-1] = matching[0]
                    prefix = "- " if action.lstrip().startswith("- ") else ""
                    repaired = prefix + " | ".join(field.lstrip("- ").strip() for field in fields)
        repaired_actions.append(repaired)
    summary["actions"] = "\n".join(repaired_actions)

    decision_evidence = [
        expression
        for source_line, expression in evidence
        if re.search(
            r"\b(decid(?:e|ed)|agree(?:d)?|confirm(?:ed)?|launch)\b|결정|합의|확정",
            source_line,
            flags=re.IGNORECASE,
        )
    ]
    if decision_evidence:
        specific_decision_dates = [
            expression
            for expression in decision_evidence
            if expression not in {"오늘", "어제"}
        ]
        decision_text = str(summary.get("decisions", "")).replace(
            unknown_date,
            (specific_decision_dates or decision_evidence)[0],
        )
        if specific_decision_dates:
            for contextual_date in ("오늘", "어제"):
                if contextual_date in decision_evidence:
                    decision_text = decision_text.replace(
                        contextual_date, specific_decision_dates[0]
                    )
        summary["decisions"] = decision_text
    return summary


def create_meeting_summary(note: dict, transcript: str) -> dict:
    instructions = (
        "당신은 한국어 업무 회의록 작성자입니다. 입력 녹취는 한국어, 영어 또는 다른 "
        "언어일 수 있습니다. 결과는 반드시 자연스러운 한국어로 작성하세요. 배경음악, "
        "노래 가사, 광고 소리와 회의와 무관한 잡담은 제외하세요. 불확실한 내용은 "
        "추측하지 말고 '확인 필요'라고 표시하세요. content에는 회의 전체 내용을 "
        "팀장 보고용 회의록처럼 정리하세요. 형식은 반드시 '1. 안건명' 다음에 빈 줄을 "
        "한 줄 두고, 각 내용을 '- 문장' 형태로 적으세요. 안건이 바뀌면 2, 3, 4 순서로 "
        "번호를 붙이세요. 문장은 짧고 명확하게 다듬되, 어느 회사나 담당자가 무엇을 "
        "검토·전달·요청하기로 했는지와 일정 및 공급가 같은 핵심 정보는 생략하지 마세요. "
        "확정 사항과 후속 업무도 별도 표나 별도 장으로 분리하지 말고 관련 안건의 "
        "글머리표 안에 자연스럽게 포함하세요. 마크다운 굵게 표시(**), 표, 담당자 | 업무 | "
        "기한 형식은 사용하지 마세요. decisions와 actions에는 빈 문자열만 반환하세요. "
        "사람 이름, 상품명, 숫자, 날짜와 "
        "요일은 녹취에 나온 표현을 그대로 보존하고 절대로 다른 값으로 바꾸지 마세요. "
        "특히 상대 날짜를 달력 날짜로 계산하지 마세요. next Monday는 '다음 월요일', "
        "by Friday는 '금요일까지', tomorrow는 '내일'처럼 표현만 번역해 보존하세요. "
        "녹취에 없는 YYYY-MM-DD 형식의 날짜를 절대로 새로 만들지 마세요. "
        "출력 형식 예시는 다음과 같습니다:\n"
        "1. 쿠킹 오일 진행 사항 및 견적 관련\n\n"
        "- 공급사에서 내부 논의 중이며, 최적의 공급가를 제안하기 위해 검토 중\n"
        "- 신규 견적을 8월 중순 이전까지 확정하여 전달해 줄 것을 요청함\n\n"
        "2. 신규 SKU 관련\n\n"
        "- 신규 SKU 정보 및 공급가표를 작성해 전달할 예정"
    )
    chunks: list[str] = []
    remaining = normalize_relative_time_for_summary(transcript)
    while len(remaining) > 50000:
        split_at = remaining.rfind("\n", 0, 50000)
        if split_at < 30000:
            split_at = 50000
        chunks.append(remaining[:split_at])
        remaining = remaining[split_at:].lstrip()
    chunks.append(remaining)
    if len(chunks) > 1:
        partials = []
        for index, chunk in enumerate(chunks, start=1):
            partial = ollama_json(
                f"{instructions}\n\n긴 회의의 {index}/{len(chunks)} 구간입니다.\n"
                f"이 구간만 중간 정리하세요.\n\n{chunk}"
            )
            partials.append(json.dumps(partial, ensure_ascii=False))
        transcript_for_summary = "\n\n".join(partials)
        source_label = "구간별 중간 정리"
    else:
        transcript_for_summary = chunks[0]
        source_label = "원문 녹취"
    prompt = (
        f"{instructions}\n\n회의 제목: {note.get('title', '')}\n"
        f"회의 날짜: {note.get('date', '')}\n"
        f"참석자: {note.get('attendees', '')}\n\n"
        f"{source_label}:\n{transcript_for_summary}"
    )
    summary = ollama_json(prompt)
    allowed_dates = set(re.findall(r"\b\d{4}-\d{2}-\d{2}\b", transcript))
    for field in ("content", "decisions", "actions"):
        value = str(summary.get(field, ""))
        summary[field] = re.sub(
            r"\b\d{4}-\d{2}-\d{2}\b",
            lambda match: (
                match.group(0)
                if match.group(0) in allowed_dates
                else "원문 날짜 확인 필요"
            ),
            value,
        )
    return restore_relative_times(summary, transcript)


def generate_meeting_notes(note_id: str) -> dict:
    status = meeting_ai_status()
    if not status["whisper_ready"]:
        raise ValueError("무료 받아쓰기 모델을 아직 내려받는 중입니다.")
    if not status["summary_ready"]:
        raise ValueError("무료 회의록 정리 모델을 아직 내려받는 중입니다.")
    with MEETING_AI_LOCK:
        with MEETING_NOTES_LOCK:
            notes = load_meeting_notes()
            note = find_meeting_note(notes, note_id)
            note_snapshot = json.loads(json.dumps(note, ensure_ascii=False))
        recordings = note_snapshot.get("recordings", [])
        if not recordings:
            raise ValueError("AI가 들을 녹음이 없습니다.")
        transcripts = []
        updated_transcripts: dict[str, tuple[str, list[dict]]] = {}
        for index, recording in enumerate(recordings, start=1):
            transcript = str(recording.get("transcript", "")).strip()
            segments = recording.get("segments", [])
            if not transcript:
                path = (
                    MEETING_RECORDINGS_DIR
                    / note_id
                    / str(recording.get("filename", ""))
                )
                if not path.is_file():
                    raise ValueError("녹음 파일을 찾을 수 없습니다.")
                transcript, segments = transcribe_recording(path)
                updated_transcripts[str(recording.get("id", ""))] = (
                    transcript,
                    segments,
                )
            transcripts.append(f"[녹음 {index}]\n{transcript}")
        transcript_text = "\n\n".join(transcripts)
        summary = create_meeting_summary(note_snapshot, transcript_text)
        content_parts = []
        main_content = str(summary.get("content", "")).strip()
        if main_content:
            content_parts.append(main_content)
        for heading, field in (
            ("결정 사항", "decisions"),
            ("후속 업무", "actions"),
        ):
            section_content = str(summary.get(field, "")).strip()
            if section_content and section_content not in {"미정", "없음"}:
                content_parts.append(f"{heading}\n{section_content}")
        with MEETING_NOTES_LOCK:
            notes = load_meeting_notes()
            note = find_meeting_note(notes, note_id)
            for recording in note.get("recordings", []):
                cached = updated_transcripts.get(str(recording.get("id", "")))
                if cached:
                    recording["transcript"], recording["segments"] = cached
            note.update(
                {
                    "content": "\n\n".join(content_parts),
                    "decisions": "",
                    "actions": "",
                    "transcript": transcript_text,
                    "ai_generated_at": datetime.now().isoformat(timespec="seconds"),
                    "ai_mode": "local_free",
                    "updated_at": datetime.now().isoformat(timespec="seconds"),
                }
            )
            save_meeting_notes(notes)
            response_note = dict(note)
    return {"ok": True, "note": response_note}


def public_meeting_ai_job(note_id: str) -> dict:
    with MEETING_AI_JOBS_LOCK:
        job = dict(MEETING_AI_JOBS.get(note_id, {}))
    if not job:
        return {
            "note_id": note_id,
            "status": "idle",
            "message": "대기 중",
        }
    job.pop("thread", None)
    return job


def _run_meeting_ai_job(note_id: str) -> None:
    with MEETING_AI_JOBS_LOCK:
        job = MEETING_AI_JOBS[note_id]
        job.update(
            {
                "status": "running",
                "message": "녹음을 받아쓰고 회의록을 작성하고 있습니다.",
                "started_at": datetime.now().isoformat(timespec="seconds"),
            }
        )
    try:
        generate_meeting_notes(note_id)
    except Exception as exc:
        with MEETING_AI_JOBS_LOCK:
            MEETING_AI_JOBS[note_id].update(
                {
                    "status": "failed",
                    "message": str(exc),
                    "finished_at": datetime.now().isoformat(timespec="seconds"),
                }
            )
    else:
        with MEETING_AI_JOBS_LOCK:
            MEETING_AI_JOBS[note_id].update(
                {
                    "status": "completed",
                    "message": "회의록 초안 작성을 완료했습니다.",
                    "finished_at": datetime.now().isoformat(timespec="seconds"),
                }
            )


def start_meeting_ai_job(note_id: str) -> dict:
    with MEETING_NOTES_LOCK:
        notes = load_meeting_notes()
        note = find_meeting_note(notes, note_id)
        if not note.get("recordings"):
            raise ValueError("AI가 들을 녹음이 없습니다.")
    with MEETING_AI_JOBS_LOCK:
        current = MEETING_AI_JOBS.get(note_id)
        if current and current.get("status") in {"queued", "running"}:
            public_current = dict(current)
            public_current.pop("thread", None)
            return {"ok": True, "started": False, "job": public_current}
        job = {
            "note_id": note_id,
            "status": "queued",
            "message": "백그라운드 작업을 시작합니다.",
            "requested_at": datetime.now().isoformat(timespec="seconds"),
        }
        MEETING_AI_JOBS[note_id] = job
        worker = threading.Thread(
            target=_run_meeting_ai_job,
            args=(note_id,),
            name=f"meeting-ai-{note_id[:8]}",
            daemon=True,
        )
        job["thread"] = worker
        worker.start()
        public_job = dict(job)
        public_job.pop("thread", None)
    return {"ok": True, "started": True, "job": public_job}


def get_meeting_notes() -> dict:
    with MEETING_NOTES_LOCK:
        notes = load_meeting_notes()
    notes.sort(
        key=lambda item: (
            str(item.get("date", "")),
            str(item.get("updated_at", "")),
        ),
        reverse=True,
    )
    return {"notes": notes}


def update_meeting_note(payload: dict) -> dict:
    action = str(payload.get("action", "")).strip()
    with MEETING_NOTES_LOCK:
        notes = load_meeting_notes()
        if action == "save":
            note_id = str(payload.get("id", "")).strip()
            title = str(payload.get("title", "")).strip()
            meeting_date = str(payload.get("date", "")).strip()
            if not title:
                raise ValueError("회의 제목을 입력해주세요.")
            if len(title) > 200:
                raise ValueError("회의 제목은 200자까지 입력할 수 있습니다.")
            try:
                date.fromisoformat(meeting_date)
            except ValueError as exc:
                raise ValueError("올바른 회의 날짜가 아닙니다.") from exc
            note = next(
                (item for item in notes if item.get("id") == note_id),
                None,
            )
            now = datetime.now().isoformat(timespec="seconds")
            if note is None:
                note = {
                    "id": uuid4().hex,
                    "created_at": now,
                }
                notes.append(note)
            note.update(
                {
                    "title": title,
                    "date": meeting_date,
                    "attendees": str(payload.get("attendees", "")).strip(),
                    "content": str(payload.get("content", "")).strip(),
                    "decisions": str(payload.get("decisions", "")).strip(),
                    "actions": str(payload.get("actions", "")).strip(),
                    "updated_at": now,
                }
            )
            response_note = dict(note)
        elif action == "delete":
            note_id = str(payload.get("id", "")).strip()
            if not any(item.get("id") == note_id for item in notes):
                raise ValueError("삭제할 회의록을 찾을 수 없습니다.")
            notes = [item for item in notes if item.get("id") != note_id]
            response_note = {"id": note_id}
        else:
            raise ValueError("지원하지 않는 회의록 작업입니다.")
        save_meeting_notes(notes)
    if action == "delete":
        shutil.rmtree(MEETING_RECORDINGS_DIR / note_id, ignore_errors=True)
    return {"ok": True, "note": response_note}


ROUTINE_LABELS = {
    "daily": "매일",
    "weekdays": "평일",
    "weekly": "매주",
    "monthly": "매월",
}
CALENDAR_STATUSES = {"todo", "doing", "review", "done"}
EVENT_TYPE_LABELS = {
    "special": "특가전",
    "group_buy": "공동구매",
    "popup": "팝업행사",
}


def calendar_status(item: dict) -> str:
    status = str(item.get("status", ""))
    if status not in CALENDAR_STATUSES:
        return "done" if item.get("completed") else "todo"
    return status


def calendar_priority(item: dict) -> int:
    try:
        priority = int(item.get("priority", 1))
    except (TypeError, ValueError):
        return 1
    return priority if priority in {1, 2, 3} else 1


def routine_occurs_on(item: dict, target: date) -> bool:
    try:
        start = date.fromisoformat(str(item.get("date", "")))
    except ValueError:
        return False
    if target < start:
        return False
    recurrence = str(item.get("recurrence", "weekly"))
    if recurrence == "daily":
        return True
    if recurrence == "weekdays":
        return target.weekday() < 5
    if recurrence == "weekly":
        return target.weekday() == start.weekday()
    if recurrence == "monthly":
        return target.day == start.day
    return False


def checklist_routine_period_key(item: dict, target: date) -> str:
    recurrence = str(item.get("recurrence", "daily"))
    if recurrence == "weekly":
        return (target - timedelta(days=target.weekday())).isoformat()
    return target.isoformat()


def get_calendar_checklist_routines() -> dict:
    target = date.today()
    with CALENDAR_LOCK:
        stored = load_calendar_tasks()
    routines = []
    for item in stored:
        if item.get("kind") != "checklist_routine":
            continue
        recurrence = str(item.get("recurrence", "daily"))
        if recurrence not in {"daily", "weekly"}:
            recurrence = "daily"
        period_key = checklist_routine_period_key(item, target)
        completed_at = dict(item.get("completions") or {}).get(period_key)
        routines.append(
            {
                "id": str(item.get("id", "")),
                "text": str(item.get("text", "")),
                "recurrence": recurrence,
                "recurrence_label": "매주" if recurrence == "weekly" else "매일",
                "period_key": period_key,
                "completed": bool(completed_at),
                "completed_at": completed_at,
                "created_at": str(item.get("created_at", "")),
            }
        )
    routines.sort(
        key=lambda item: (
            bool(item.get("completed")),
            str(item.get("created_at", "")),
        )
    )
    return {"date": target.isoformat(), "routines": routines}


def update_calendar_checklist_routine(payload: dict) -> dict:
    action = str(payload.get("action", "")).strip()
    now = datetime.now().isoformat(timespec="seconds")
    target = date.today()
    with CALENDAR_LOCK:
        tasks = load_calendar_tasks()
        if action == "add":
            text = str(payload.get("text", "")).strip()
            recurrence = str(payload.get("recurrence", "daily")).strip()
            if not text:
                raise ValueError("루틴 업무를 입력해주세요.")
            if len(text) > 200:
                raise ValueError("루틴 업무는 200자까지 입력할 수 있습니다.")
            if recurrence not in {"daily", "weekly"}:
                raise ValueError("루틴 반복 주기는 매일 또는 매주만 선택할 수 있습니다.")
            routine = {
                "id": uuid4().hex,
                "kind": "checklist_routine",
                "text": text,
                "recurrence": recurrence,
                "completions": {},
                "created_at": now,
                "updated_at": now,
            }
            tasks.append(routine)
            result = dict(routine)
            result.update(
                {
                    "recurrence_label": "매주" if recurrence == "weekly" else "매일",
                    "period_key": checklist_routine_period_key(routine, target),
                    "completed": False,
                    "completed_at": None,
                }
            )
        elif action == "toggle":
            routine_id = str(payload.get("id", "")).strip()
            routine = next(
                (
                    item
                    for item in tasks
                    if item.get("id") == routine_id
                    and item.get("kind") == "checklist_routine"
                ),
                None,
            )
            if routine is None:
                raise ValueError("루틴 업무를 찾을 수 없습니다.")
            period_key = checklist_routine_period_key(routine, target)
            completions = dict(routine.get("completions") or {})
            if bool(payload.get("completed", False)):
                completions[period_key] = now
            else:
                completions.pop(period_key, None)
            routine["completions"] = completions
            routine["updated_at"] = now
            result = {
                **routine,
                "recurrence_label": (
                    "매주" if routine.get("recurrence") == "weekly" else "매일"
                ),
                "period_key": period_key,
                "completed": period_key in completions,
                "completed_at": completions.get(period_key),
            }
        elif action == "delete":
            routine_id = str(payload.get("id", "")).strip()
            if not any(
                item.get("id") == routine_id
                and item.get("kind") == "checklist_routine"
                for item in tasks
            ):
                raise ValueError("삭제할 루틴 업무를 찾을 수 없습니다.")
            tasks = [item for item in tasks if item.get("id") != routine_id]
            result = {"id": routine_id}
        else:
            raise ValueError("지원하지 않는 루틴 작업입니다.")
        save_calendar_tasks(tasks)
    return {"ok": True, "routine": result}


def routine_occurrence(item: dict, target: date) -> dict:
    task_date = target.isoformat()
    completed_at = dict(item.get("completions") or {}).get(task_date)
    status = dict(item.get("statuses") or {}).get(task_date)
    if status not in CALENDAR_STATUSES:
        status = "done" if completed_at else "todo"
    recurrence = str(item.get("recurrence", "weekly"))
    return {
        "id": f"{item['id']}:{task_date}",
        "source_id": item["id"],
        "source_date": str(item.get("date", "")),
        "date": task_date,
        "text": str(item.get("text", "")),
        "kind": "routine",
        "recurrence": recurrence,
        "recurrence_label": ROUTINE_LABELS.get(recurrence, "반복"),
        "priority": calendar_priority(item),
        "status": status,
        "completed": status == "done",
        "created_at": str(item.get("created_at", "")),
        "completed_at": completed_at,
    }


def prepare_routine_today(item: dict, target: date) -> tuple[dict | None, bool]:
    """Keep the routine shown each day until its current occurrence is done."""
    target_key = target.isoformat()
    statuses = dict(item.get("statuses") or {})
    completions = dict(item.get("completions") or {})
    rollover_key = str(item.get("rollover_date", "")).strip()
    changed = False

    try:
        rollover_day = date.fromisoformat(rollover_key) if rollover_key else None
    except ValueError:
        rollover_day = None
        item.pop("rollover_date", None)
        changed = True

    if rollover_day and rollover_day <= target:
        rollover_status = statuses.get(rollover_key)
        if rollover_status not in CALENDAR_STATUSES:
            rollover_status = "done" if completions.get(rollover_key) else "todo"
        if rollover_status == "done" and rollover_day < target:
            item.pop("rollover_date", None)
            rollover_day = None
            rollover_key = ""
            changed = True
        elif rollover_status != "done" and rollover_day < target:
            current_status = statuses.get(target_key)
            if current_status not in CALENDAR_STATUSES:
                current_status = (
                    "done" if completions.get(target_key) else rollover_status
                )
                statuses[target_key] = current_status
                item["statuses"] = statuses
            item["rollover_date"] = target_key
            rollover_day = target
            rollover_key = target_key
            changed = True

    occurs_today = routine_occurs_on(item, target)
    if rollover_day is None and occurs_today:
        current_status = statuses.get(target_key)
        if current_status not in CALENDAR_STATUSES:
            current_status = "done" if completions.get(target_key) else "todo"
        if current_status != "done":
            item["rollover_date"] = target_key
            rollover_day = target
            changed = True

    if occurs_today or rollover_day == target:
        return routine_occurrence(item, target), changed
    return None, changed


def event_occurs_on(item: dict, target: date) -> bool:
    try:
        start = date.fromisoformat(str(item.get("date", "")))
        end = date.fromisoformat(str(item.get("end_date", "")))
    except ValueError:
        return False
    return start <= target <= end


def event_occurrence(item: dict, target: date) -> dict:
    task_date = target.isoformat()
    completed_at = dict(item.get("completions") or {}).get(task_date)
    status = dict(item.get("statuses") or {}).get(task_date)
    if status not in CALENDAR_STATUSES:
        status = "done" if completed_at else "todo"
    start = date.fromisoformat(str(item["date"]))
    end = date.fromisoformat(str(item["end_date"]))
    event_label = (
        f"{start.month}/{start.day}"
        if start == end
        else f"{start.month}/{start.day}~{end.month}/{end.day}"
    )
    event_type = str(item.get("event_type", "special"))
    if event_type not in EVENT_TYPE_LABELS:
        event_type = "special"
    return {
        "id": f"{item['id']}:{task_date}",
        "source_id": item["id"],
        "date": task_date,
        "text": str(item.get("text", "")),
        "kind": "event",
        "event_start": start.isoformat(),
        "event_end": end.isoformat(),
        "event_label": event_label,
        "event_type": event_type,
        "event_type_label": EVENT_TYPE_LABELS[event_type],
        "source": str(item.get("source", "")),
        "source_spreadsheet_id": str(item.get("source_spreadsheet_id", "")),
        "status": status,
        "completed": status == "done",
        "created_at": str(item.get("created_at", "")),
        "completed_at": completed_at,
    }


def source_occurs_on(item: dict, target: date) -> bool:
    if item.get("kind") == "routine":
        return routine_occurs_on(item, target)
    if item.get("kind") == "event":
        return event_occurs_on(item, target)
    return False


def source_occurrence(item: dict, target: date) -> dict:
    if item.get("kind") == "event":
        return event_occurrence(item, target)
    return routine_occurrence(item, target)


def source_can_update_on(item: dict, target: date) -> bool:
    if source_occurs_on(item, target):
        return True
    return (
        item.get("kind") == "routine"
        and str(item.get("rollover_date", "")) == target.isoformat()
    )


def get_calendar_month(month: str) -> dict:
    try:
        month_start = datetime.strptime(f"{month}-01", "%Y-%m-%d").date()
    except ValueError as exc:
        raise ValueError("올바른 월 형식이 아닙니다.") from exc
    if month_start.month == 12:
        next_month = date(month_start.year + 1, 1, 1)
    else:
        next_month = date(month_start.year, month_start.month + 1, 1)
    with CALENDAR_LOCK:
        stored = load_calendar_tasks()
    tasks: list[dict] = []
    for item in stored:
        if item.get("kind") == "checklist_routine":
            continue
        if item.get("kind") in {"routine", "event"}:
            target = month_start
            target_end = next_month
            while target < target_end:
                if source_occurs_on(item, target):
                    tasks.append(source_occurrence(item, target))
                target += timedelta(days=1)
        elif str(item.get("date", "")).startswith(f"{month}-"):
            task = dict(item)
            task.setdefault("kind", "task")
            status = calendar_status(task)
            task["status"] = status
            task["completed"] = status == "done"
            task["priority"] = calendar_priority(task)
            tasks.append(task)
    tasks.sort(
        key=lambda item: (
            str(item.get("date", "")),
            bool(item.get("completed", False)),
            str(item.get("created_at", "")),
        )
    )
    return {"month": month, "tasks": tasks}


def get_calendar_today() -> dict:
    target = date.today()
    target_key = target.isoformat()
    changed = False
    with CALENDAR_LOCK:
        stored = load_calendar_tasks()
        for item in stored:
            if item.get("kind") == "checklist_routine":
                continue
            if item.get("kind") == "routine":
                _, routine_changed = prepare_routine_today(item, target)
                changed = changed or routine_changed
                continue
            if item.get("kind") == "event":
                continue
            try:
                scheduled = date.fromisoformat(str(item.get("date", "")))
            except ValueError:
                continue
            status = calendar_status(item)
            if scheduled < target and status != "done":
                item["date"] = target_key
                item["status"] = status
                item["completed"] = False
                changed = True
        if changed:
            save_calendar_tasks(stored)

    tasks: list[dict] = []
    nearby_events: list[dict] = []
    nearby_start = target - timedelta(days=1)
    nearby_end = target + timedelta(days=2)
    for item in stored:
        if item.get("kind") == "checklist_routine":
            continue
        if item.get("kind") == "event":
            try:
                event_start = date.fromisoformat(str(item.get("date", "")))
                event_end = date.fromisoformat(
                    str(item.get("end_date", ""))
                )
            except ValueError:
                continue
            if event_start <= nearby_end and event_end >= nearby_start:
                occurrence_date = max(event_start, nearby_start)
                nearby_events.append(event_occurrence(item, occurrence_date))
        elif item.get("kind") == "routine":
            occurrence, _ = prepare_routine_today(item, target)
            if occurrence:
                tasks.append(occurrence)
        elif str(item.get("date", "")) == target_key:
            task = dict(item)
            task.setdefault("kind", "task")
            status = calendar_status(task)
            task["status"] = status
            task["completed"] = status == "done"
            task["priority"] = calendar_priority(task)
            tasks.append(task)
    tasks.sort(
        key=lambda item: (
            bool(item.get("completed", False)),
            str(item.get("created_at", "")),
        )
    )
    nearby_events.sort(
        key=lambda item: (
            str(item.get("event_start", "")),
            str(item.get("event_end", "")),
            str(item.get("text", "")),
        )
    )
    return {
        "date": target_key,
        "tasks": tasks,
        "nearby_start": nearby_start.isoformat(),
        "nearby_end": nearby_end.isoformat(),
        "nearby_events": nearby_events,
    }


def sync_calendar_events() -> dict:
    sync_config = dict(CONFIG.get("calendar_sync") or {})
    if not sync_config:
        raise RuntimeError("행사 일정 동기화 설정이 없습니다.")
    if not CALENDAR_SYNC_LOCK.acquire(blocking=False):
        raise RuntimeError("행사 일정 동기화가 이미 진행 중입니다.")
    CALENDAR_SYNC_STATUS.update(
        status="running",
        message="Google Sheets 행사 일정을 확인하고 있습니다.",
    )
    try:
        sheet_events = fetch_sheet_events(APP_DIR, sync_config)
        with CALENDAR_LOCK:
            tasks = load_calendar_tasks()
            merged, result = merge_sheet_events(
                tasks,
                sheet_events,
                str(sync_config["spreadsheet_id"]),
            )
            save_calendar_tasks(merged)
        CALENDAR_SYNC_STATUS.update(
            status="completed",
            message=f"행사 일정 {result['total']}개를 최신화했습니다.",
            last_synced_at=result["synced_at"],
        )
        return {
            "ok": True,
            **result,
            "message": CALENDAR_SYNC_STATUS["message"],
        }
    except Exception as exc:
        CALENDAR_SYNC_STATUS.update(
            status="error",
            message=str(exc),
        )
        raise
    finally:
        CALENDAR_SYNC_LOCK.release()


def update_calendar_task(payload: dict) -> dict:
    action = str(payload.get("action", "")).strip()
    with CALENDAR_LOCK:
        tasks = load_calendar_tasks()
        if action == "add":
            task_date = str(payload.get("date", "")).strip()
            text = str(payload.get("text", "")).strip()
            kind = str(payload.get("kind", "task")).strip()
            try:
                priority = int(payload.get("priority", 1))
            except (TypeError, ValueError) as exc:
                raise ValueError("올바른 우선순위가 아닙니다.") from exc
            try:
                parsed_date = date.fromisoformat(task_date)
            except ValueError as exc:
                raise ValueError("올바른 날짜가 아닙니다.") from exc
            if not text:
                raise ValueError("업무 내용을 입력해주세요.")
            if len(text) > 200:
                raise ValueError("업무 내용은 200자까지 입력할 수 있습니다.")
            if kind not in {"task", "meeting", "routine", "event"}:
                raise ValueError("올바른 업무 유형이 아닙니다.")
            if priority not in {1, 2, 3}:
                raise ValueError("우선순위는 별 1개부터 3개까지 선택해주세요.")
            task = {
                "id": uuid4().hex,
                "date": task_date,
                "text": text,
                "kind": kind,
                "status": "todo",
                "priority": priority,
                "completed": False,
                "created_at": datetime.now().isoformat(timespec="seconds"),
                "completed_at": None,
            }
            if kind == "routine":
                recurrence = str(payload.get("recurrence", "weekly")).strip()
                if recurrence not in ROUTINE_LABELS:
                    raise ValueError("올바른 반복 방식이 아닙니다.")
                task["recurrence"] = recurrence
                task["completions"] = {}
                task["statuses"] = {}
            elif kind == "event":
                event_type = str(payload.get("event_type", "special")).strip()
                if event_type not in EVENT_TYPE_LABELS:
                    raise ValueError("올바른 행사 유형이 아닙니다.")
                try:
                    end_date = date.fromisoformat(
                        str(payload.get("end_date", "")).strip()
                    )
                except ValueError as exc:
                    raise ValueError("행사 종료일을 선택해주세요.") from exc
                if end_date < parsed_date:
                    raise ValueError("행사 종료일은 시작일보다 빠를 수 없습니다.")
                task["end_date"] = end_date.isoformat()
                task["event_type"] = event_type
                task["completions"] = {}
                task["statuses"] = {}
            tasks.append(task)
            response_task = source_occurrence(task, parsed_date) if kind in {
                "routine",
                "event",
            } else task
        elif action == "edit":
            task_id = str(payload.get("id", "")).strip()
            source_id = str(payload.get("source_id", "")).strip()
            edit_id = source_id or task_id
            task = next(
                (item for item in tasks if item.get("id") == edit_id),
                None,
            )
            if task is None:
                raise ValueError("수정할 일정을 찾을 수 없습니다.")
            task_date = str(payload.get("date", "")).strip()
            text = str(payload.get("text", "")).strip()
            kind = str(payload.get("kind", "task")).strip()
            try:
                parsed_date = date.fromisoformat(task_date)
            except ValueError as exc:
                raise ValueError("올바른 날짜가 아닙니다.") from exc
            try:
                priority = int(payload.get("priority", 1))
            except (TypeError, ValueError) as exc:
                raise ValueError("올바른 우선순위가 아닙니다.") from exc
            if not text:
                raise ValueError("업무 내용을 입력해주세요.")
            if len(text) > 200:
                raise ValueError("업무 내용은 200자까지 입력할 수 있습니다.")
            if kind not in {"task", "meeting", "routine", "event"}:
                raise ValueError("올바른 업무 유형이 아닙니다.")
            if priority not in {1, 2, 3}:
                raise ValueError("우선순위는 별 1개부터 3개까지 선택해주세요.")

            previous_kind = str(task.get("kind", "task"))
            task.update(
                {
                    "date": task_date,
                    "text": text,
                    "kind": kind,
                    "priority": priority,
                    "updated_at": datetime.now().isoformat(timespec="seconds"),
                }
            )
            if kind == "routine":
                recurrence = str(payload.get("recurrence", "weekly")).strip()
                if recurrence not in ROUTINE_LABELS:
                    raise ValueError("올바른 반복 방식이 아닙니다.")
                task["recurrence"] = recurrence
                task.setdefault("completions", {})
                task.setdefault("statuses", {})
                task.pop("end_date", None)
                task.pop("event_type", None)
            elif kind == "event":
                event_type = str(payload.get("event_type", "special")).strip()
                if event_type not in EVENT_TYPE_LABELS:
                    raise ValueError("올바른 행사 유형이 아닙니다.")
                try:
                    end_date = date.fromisoformat(
                        str(payload.get("end_date", "")).strip()
                    )
                except ValueError as exc:
                    raise ValueError("행사 종료일을 선택해주세요.") from exc
                if end_date < parsed_date:
                    raise ValueError("행사 종료일은 시작일보다 빠를 수 없습니다.")
                task["end_date"] = end_date.isoformat()
                task["event_type"] = event_type
                task.setdefault("completions", {})
                task.setdefault("statuses", {})
                task.pop("recurrence", None)
            else:
                task.pop("recurrence", None)
                task.pop("end_date", None)
                task.pop("event_type", None)
                task.pop("completions", None)
                task.pop("statuses", None)
                if previous_kind in {"routine", "event"}:
                    task["status"] = "todo"
                    task["completed"] = False
                    task["completed_at"] = None

            response_task = (
                source_occurrence(task, parsed_date)
                if kind in {"routine", "event"}
                else task
            )
        elif action == "toggle":
            task_id = str(payload.get("id", "")).strip()
            source_id = str(payload.get("source_id", "")).strip()
            if source_id:
                task = next(
                    (
                        item
                        for item in tasks
                        if item.get("id") == source_id
                        and item.get("kind") in {"routine", "event"}
                    ),
                    None,
                )
                if task is None:
                    raise ValueError("반복 또는 행사 일정을 찾을 수 없습니다.")
                try:
                    target_date = date.fromisoformat(
                        str(payload.get("date", ""))
                    )
                except ValueError as exc:
                    raise ValueError("올바른 일정 날짜가 아닙니다.") from exc
                if not source_can_update_on(task, target_date):
                    raise ValueError("해당 날짜에 표시되는 일정이 아닙니다.")
                completions = dict(task.get("completions") or {})
                statuses = dict(task.get("statuses") or {})
                if bool(payload.get("completed", False)):
                    completions[target_date.isoformat()] = (
                        datetime.now().isoformat(timespec="seconds")
                    )
                    statuses[target_date.isoformat()] = "done"
                else:
                    completions.pop(target_date.isoformat(), None)
                    statuses[target_date.isoformat()] = "todo"
                task["completions"] = completions
                task["statuses"] = statuses
                response_task = source_occurrence(task, target_date)
            else:
                task = next(
                    (item for item in tasks if item.get("id") == task_id),
                    None,
                )
                if task is None:
                    raise ValueError("업무를 찾을 수 없습니다.")
                task["completed"] = bool(payload.get("completed", False))
                task["status"] = "done" if task["completed"] else "todo"
                task["completed_at"] = (
                    datetime.now().isoformat(timespec="seconds")
                    if task["completed"]
                    else None
                )
                response_task = task
        elif action == "status":
            task_id = str(payload.get("id", "")).strip()
            source_id = str(payload.get("source_id", "")).strip()
            status = str(payload.get("status", "")).strip()
            if status not in CALENDAR_STATUSES:
                raise ValueError("올바른 업무 상태가 아닙니다.")
            if source_id:
                task = next(
                    (
                        item
                        for item in tasks
                        if item.get("id") == source_id
                        and item.get("kind") in {"routine", "event"}
                    ),
                    None,
                )
                if task is None:
                    raise ValueError("반복 또는 행사 일정을 찾을 수 없습니다.")
                try:
                    target_date = date.fromisoformat(
                        str(payload.get("date", ""))
                    )
                except ValueError as exc:
                    raise ValueError("올바른 일정 날짜가 아닙니다.") from exc
                if not source_can_update_on(task, target_date):
                    raise ValueError("해당 날짜에 표시되는 일정이 아닙니다.")
                statuses = dict(task.get("statuses") or {})
                statuses[target_date.isoformat()] = status
                task["statuses"] = statuses
                completions = dict(task.get("completions") or {})
                if status == "done":
                    completions[target_date.isoformat()] = (
                        datetime.now().isoformat(timespec="seconds")
                    )
                else:
                    completions.pop(target_date.isoformat(), None)
                task["completions"] = completions
                response_task = source_occurrence(task, target_date)
            else:
                task = next(
                    (item for item in tasks if item.get("id") == task_id),
                    None,
                )
                if task is None:
                    raise ValueError("업무를 찾을 수 없습니다.")
                task["status"] = status
                task["completed"] = status == "done"
                task["completed_at"] = (
                    datetime.now().isoformat(timespec="seconds")
                    if status == "done"
                    else None
                )
                response_task = task
        elif action == "priority":
            task_id = str(payload.get("id", "")).strip()
            source_id = str(payload.get("source_id", "")).strip()
            try:
                priority = int(payload.get("priority", 1))
            except (TypeError, ValueError) as exc:
                raise ValueError("올바른 우선순위가 아닙니다.") from exc
            if priority not in {1, 2, 3}:
                raise ValueError("우선순위는 별 1개부터 3개까지 선택해주세요.")
            if source_id:
                task = next(
                    (
                        item
                        for item in tasks
                        if item.get("id") == source_id
                        and item.get("kind") in {"routine", "event"}
                    ),
                    None,
                )
                if task is None:
                    raise ValueError("반복 또는 행사 일정을 찾을 수 없습니다.")
                try:
                    target_date = date.fromisoformat(
                        str(payload.get("date", ""))
                    )
                except ValueError as exc:
                    raise ValueError("올바른 일정 날짜가 아닙니다.") from exc
                if not source_can_update_on(task, target_date):
                    raise ValueError("해당 날짜에 표시되는 일정이 아닙니다.")
                task["priority"] = priority
                response_task = source_occurrence(task, target_date)
            else:
                task = next(
                    (item for item in tasks if item.get("id") == task_id),
                    None,
                )
                if task is None:
                    raise ValueError("업무를 찾을 수 없습니다.")
                task["priority"] = priority
                response_task = task
        elif action == "delete":
            task_id = str(payload.get("id", "")).strip()
            source_id = str(payload.get("source_id", "")).strip()
            delete_id = source_id or task_id
            task = next(
                (item for item in tasks if item.get("id") == delete_id),
                None,
            )
            if task is None:
                raise ValueError("삭제할 일정을 찾을 수 없습니다.")
            tasks = [
                item for item in tasks if item.get("id") != delete_id
            ]
            deleted_scope = "routine" if source_id else "task"
        else:
            raise ValueError("지원하지 않는 캘린더 작업입니다.")
        save_calendar_tasks(tasks)
    if action == "delete":
        return {
            "ok": True,
            "deleted_id": delete_id,
            "deleted_scope": deleted_scope,
        }
    return {"ok": True, "task": response_task}


def get_brand(brand_key: str) -> dict:
    if brand_key not in CONFIG["brands"]:
        raise ValueError(f"지원하지 않는 브랜드입니다: {brand_key}")
    return CONFIG["brands"][brand_key]


def normalize_instagram_id(value: object) -> str:
    text = str(value or "").strip()
    if text.startswith("@"):
        text = text[1:]
    return text.strip("/")


def normalize_profile_url(value: object, instagram_id: str) -> str:
    text = str(value or "").strip()
    if text.startswith("http://") or text.startswith("https://"):
        return text
    return f"https://www.instagram.com/{instagram_id}/" if instagram_id else ""


def parse_sheet_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y.%m.%d", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def load_dm_template_overrides() -> dict[str, str]:
    with DM_TEMPLATE_LOCK:
        if not DM_TEMPLATE_OVERRIDES_PATH.exists():
            return {}
        try:
            raw = json.loads(
                DM_TEMPLATE_OVERRIDES_PATH.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError):
            return {}
        if not isinstance(raw, dict):
            return {}
        return {
            str(key): str(value)
            for key, value in raw.items()
            if isinstance(value, str)
        }


def get_dm_message_template(brand_key: str, brand: dict) -> str:
    overrides = load_dm_template_overrides()
    return overrides.get(brand_key, str(brand["message_template"]))


def save_dm_message_template(brand_key: str, template: str) -> dict:
    if brand_key not in CONFIG["brands"]:
        raise ValueError("지원하지 않는 브랜드입니다.")
    if not isinstance(template, str):
        raise ValueError("DM 문구를 입력해주세요.")
    template = template.strip()
    if not template:
        raise ValueError("DM 문구를 비워둘 수 없습니다.")
    if len(template) > 20000:
        raise ValueError("DM 문구는 20,000자 이내로 저장해주세요.")
    with DM_TEMPLATE_LOCK:
        overrides: dict[str, str] = {}
        if DM_TEMPLATE_OVERRIDES_PATH.exists():
            try:
                raw = json.loads(
                    DM_TEMPLATE_OVERRIDES_PATH.read_text(encoding="utf-8")
                )
                if isinstance(raw, dict):
                    overrides = {
                        str(key): str(value)
                        for key, value in raw.items()
                        if isinstance(value, str)
                    }
            except (OSError, json.JSONDecodeError):
                pass
        overrides[brand_key] = template
        DM_TEMPLATE_OVERRIDES_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary = DM_TEMPLATE_OVERRIDES_PATH.with_suffix(".tmp")
        temporary.write_text(
            json.dumps(overrides, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        temporary.replace(DM_TEMPLATE_OVERRIDES_PATH)
    return {"ok": True, "brand": brand_key, "template": template}


def reset_dm_message_template(brand_key: str) -> dict:
    if brand_key not in CONFIG["brands"]:
        raise ValueError("지원하지 않는 브랜드입니다.")
    with DM_TEMPLATE_LOCK:
        overrides: dict[str, str] = {}
        if DM_TEMPLATE_OVERRIDES_PATH.exists():
            try:
                raw = json.loads(
                    DM_TEMPLATE_OVERRIDES_PATH.read_text(encoding="utf-8")
                )
                if isinstance(raw, dict):
                    overrides = {
                        str(key): str(value)
                        for key, value in raw.items()
                        if isinstance(value, str)
                    }
            except (OSError, json.JSONDecodeError):
                pass
        overrides.pop(brand_key, None)
        DM_TEMPLATE_OVERRIDES_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary = DM_TEMPLATE_OVERRIDES_PATH.with_suffix(".tmp")
        temporary.write_text(
            json.dumps(overrides, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        temporary.replace(DM_TEMPLATE_OVERRIDES_PATH)
    template = str(CONFIG["brands"][brand_key]["message_template"])
    return {"ok": True, "brand": brand_key, "template": template}


def format_message(template: str, instagram_id: str) -> str:
    if template == "__MESSAGE_REQUIRED__":
        return ""
    return template.replace("{instagram_id}", instagram_id)


def open_sheet(brand: dict):
    path = CONFIG["_workbook_path"]
    if not path.exists():
        raise FileNotFoundError(f"작업용 Excel 파일을 찾을 수 없습니다: {path}")
    workbook = load_workbook(path, data_only=False, keep_links=True)
    sheet_name = brand["sheet_name"]
    if sheet_name not in workbook.sheetnames:
        raise KeyError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
    return workbook, workbook[sheet_name]


def header_map(sheet) -> dict[str, int]:
    return {
        str(sheet.cell(1, column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
    }


def required_columns(sheet, brand: dict) -> dict[str, int]:
    headers = header_map(sheet)
    resolved = {}
    for key, header in brand["columns"].items():
        if header not in headers:
            raise KeyError(f"필수 열 '{header}'을(를) 찾을 수 없습니다.")
        resolved[key] = headers[header]
    return resolved


def load_skip_log() -> dict:
    if not SKIP_LOG_PATH.exists():
        return {}
    with SKIP_LOG_PATH.open("r", encoding="utf-8") as file:
        return json.load(file)


def save_skip_log(skip_log: dict) -> None:
    temporary = SKIP_LOG_PATH.with_suffix(".saving.json")
    with temporary.open("w", encoding="utf-8") as file:
        json.dump(skip_log, file, ensure_ascii=False, indent=2)
    os.replace(temporary, SKIP_LOG_PATH)


def load_dm_sync_state() -> dict:
    if not DM_SYNC_STATE_PATH.exists():
        return {}
    try:
        return json.loads(DM_SYNC_STATE_PATH.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return {}


def save_dm_sync_state(state: dict) -> None:
    DM_SYNC_STATE_PATH.parent.mkdir(parents=True, exist_ok=True)
    temporary = DM_SYNC_STATE_PATH.with_suffix(".tmp")
    temporary.write_text(
        json.dumps(state, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.replace(temporary, DM_SYNC_STATE_PATH)


def load_dm_drive_service():
    from google.oauth2.credentials import Credentials
    from googleapiclient.discovery import build

    token_info = json.loads(DM_DRIVE_TOKEN_PATH.read_text(encoding="utf-8"))
    refresh_token = str(token_info.get("refresh_token", "")).strip()
    if not refresh_token:
        raise RuntimeError("Google Drive refresh token is missing.")
    refresh_payload = urlencode(
        {
            "grant_type": "refresh_token",
            "refresh_token": refresh_token,
            "client_id": token_info.get("client_id", ""),
            "client_secret": token_info.get("client_secret", ""),
        }
    ).encode("utf-8")
    refresh_request = urllib.request.Request(
        str(token_info.get("token_uri", "https://oauth2.googleapis.com/token")),
        data=refresh_payload,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    with urllib.request.urlopen(refresh_request, timeout=30) as response:
        refresh_result = json.loads(response.read().decode("utf-8"))
    access_token = str(refresh_result.get("access_token", "")).strip()
    if not access_token:
        raise RuntimeError("Google Drive access token refresh failed.")
    token_info["token"] = access_token
    DM_DRIVE_TOKEN_PATH.write_text(
        json.dumps(token_info, ensure_ascii=False),
        encoding="utf-8",
    )
    credentials = Credentials(token=access_token)
    return build("drive", "v3", credentials=credentials, cache_discovery=False)


def download_dm_workbook(temporary: Path, download_url: str) -> str:
    sync_config = dict(CONFIG.get("dm_sync") or {})
    drive_file_id = str(sync_config.get("drive_file_id", "")).strip()
    if drive_file_id and DM_DRIVE_TOKEN_PATH.exists():
        from googleapiclient.http import MediaIoBaseDownload

        service = load_dm_drive_service()
        metadata = (
            service.files()
            .get(fileId=drive_file_id, fields="modifiedTime,size")
            .execute()
        )
        if int(metadata.get("size", "0") or 0) > 25 * 1024 * 1024:
            raise ValueError("The Drive workbook is larger than the allowed size.")
        request = service.files().get_media(fileId=drive_file_id)
        with temporary.open("wb") as output:
            downloader = MediaIoBaseDownload(output, request)
            done = False
            while not done:
                _, done = downloader.next_chunk()
        return str(metadata.get("modifiedTime", ""))

    head_request = urllib.request.Request(
        download_url,
        headers={"User-Agent": "TobusWorkMaster/1.0"},
        method="HEAD",
    )
    with urllib.request.urlopen(head_request, timeout=30) as response:
        remote_modified = response.headers.get("Last-Modified", "")
    download_request = urllib.request.Request(
        download_url,
        headers={"User-Agent": "TobusWorkMaster/1.0"},
    )
    with urllib.request.urlopen(download_request, timeout=60) as response:
        content_length = int(response.headers.get("Content-Length", "0") or 0)
        if content_length > 25 * 1024 * 1024:
            raise ValueError("The remote workbook is larger than the allowed size.")
        with temporary.open("wb") as output:
            shutil.copyfileobj(response, output)
    return remote_modified


def merge_local_dm_records(local_path: Path, downloaded_path: Path) -> bool:
    """온라인 최신 명단에 앱에서 기록한 발송일을 크리에이터 ID 기준으로 보존한다."""
    if not local_path.exists():
        return False
    local_book = load_workbook(local_path, data_only=False, keep_links=True)
    remote_book = load_workbook(downloaded_path, data_only=False, keep_links=True)
    changed = False
    try:
        for brand in CONFIG["brands"].values():
            sheet_name = brand["sheet_name"]
            if sheet_name not in local_book.sheetnames or sheet_name not in remote_book.sheetnames:
                continue
            local_sheet = local_book[sheet_name]
            remote_sheet = remote_book[sheet_name]
            local_columns = required_columns(local_sheet, brand)
            remote_columns = required_columns(remote_sheet, brand)
            sent_by_creator = {}
            for row in range(2, local_sheet.max_row + 1):
                creator = normalize_instagram_id(
                    local_sheet.cell(row, local_columns["creator"]).value
                )
                sent_value = local_sheet.cell(row, local_columns["sent_date"]).value
                if creator and sent_value:
                    sent_by_creator[creator] = sent_value
            for row in range(2, remote_sheet.max_row + 1):
                creator = normalize_instagram_id(
                    remote_sheet.cell(row, remote_columns["creator"]).value
                )
                sent_cell = remote_sheet.cell(row, remote_columns["sent_date"])
                if creator in sent_by_creator and not sent_cell.value:
                    sent_cell.value = sent_by_creator[creator]
                    changed = True
        if changed:
            remote_book.save(downloaded_path)
        return changed
    finally:
        local_book.close()
        remote_book.close()


def sync_dm_workbook(force: bool = False) -> dict:
    sync_config = CONFIG.get("dm_sync", {})
    download_url = str(sync_config.get("download_url", "")).strip()
    if not download_url:
        return dict(DM_SYNC_STATUS)

    with DM_SYNC_LOCK:
        now = datetime.now()
        interval = int(sync_config.get("check_interval_seconds", 60))
        state = load_dm_sync_state()
        last_checked = (
            DM_SYNC_STATUS.get("last_checked_at")
            or state.get("last_synced_at")
        )
        if not force and last_checked:
            try:
                checked_at = datetime.fromisoformat(last_checked)
                if (now - checked_at).total_seconds() < interval:
                    DM_SYNC_STATUS.update(
                        status="current",
                        message="온라인 최신 명단을 사용 중입니다.",
                        last_checked_at=last_checked,
                        last_synced_at=state.get("last_synced_at"),
                        last_modified=state.get("last_modified"),
                    )
                    return dict(DM_SYNC_STATUS)
            except ValueError:
                pass

        DM_SYNC_STATUS.update(
            status="checking",
            message="온라인 최신 명단을 확인하고 있습니다.",
            last_checked_at=now.isoformat(timespec="seconds"),
        )
        try:
            destination = CONFIG["_workbook_path"]
            temporary = destination.with_name(
                f".{destination.stem}.downloading{destination.suffix}"
            )
            remote_modified = download_dm_workbook(temporary, download_url)
            if (
                not force
                and remote_modified
                and remote_modified == state.get("last_modified")
                and CONFIG["_workbook_path"].exists()
            ):
                DM_SYNC_STATUS.update(
                    status="current",
                    message="온라인 최신 명단과 동일합니다.",
                    last_checked_at=datetime.now().isoformat(timespec="seconds"),
                    last_modified=remote_modified,
                )
                return dict(DM_SYNC_STATUS)

            # 파일이 정상 Excel인지 확인하고 앱에서 기록한 발송일을 보존한다.
            validation_book = load_workbook(temporary, read_only=True, data_only=False)
            validation_book.close()
            with LOCK:
                merge_local_dm_records(destination, temporary)
                previous = destination.with_name(
                    f".{destination.stem}.before_sync{destination.suffix}"
                )
                if destination.exists():
                    shutil.copy2(destination, previous)
                os.replace(temporary, destination)

            synced_at = datetime.now().isoformat(timespec="seconds")
            state = {
                "last_modified": remote_modified,
                "last_synced_at": synced_at,
            }
            save_dm_sync_state(state)
            DM_SYNC_STATUS.update(
                status="completed",
                message="온라인 최신 명단을 반영했습니다.",
                last_checked_at=synced_at,
                last_synced_at=synced_at,
                last_modified=remote_modified,
            )
        except Exception as exc:
            DM_SYNC_STATUS.update(
                status="error",
                message=f"온라인 최신화 실패: {exc}",
                last_checked_at=datetime.now().isoformat(timespec="seconds"),
            )
        finally:
            try:
                if "temporary" in locals() and temporary.exists():
                    temporary.unlink()
            except OSError:
                pass
        return dict(DM_SYNC_STATUS)


def get_dashboard(brand_key: str, force_sync: bool = False) -> dict:
    brand = get_brand(brand_key)
    sync_status = sync_dm_workbook(force=force_sync)
    with LOCK:
        workbook, sheet = open_sheet(brand)
        try:
            columns = required_columns(sheet, brand)
            today = date.today()
            week_start = today - timedelta(days=today.weekday())
            week_end = week_start + timedelta(days=6)
            weekly_sent = 0
            pending = []
            skip_log = load_skip_log()
            excluded_topics = set(brand.get("excluded_topics", []))
            excluded_topic_count = 0

            for row in range(2, sheet.max_row + 1):
                creator = normalize_instagram_id(sheet.cell(row, columns["creator"]).value)
                topic = str(sheet.cell(row, columns["topic"]).value or "").strip()
                profile_url = normalize_profile_url(
                    sheet.cell(row, columns["profile_url"]).value, creator
                )
                sent_value = sheet.cell(row, columns["sent_date"]).value
                sent_date = parse_sheet_date(sent_value)
                is_excluded_topic = topic in excluded_topics

                if is_excluded_topic:
                    excluded_topic_count += 1
                if sent_date and not is_excluded_topic and week_start <= sent_date <= week_end:
                    weekly_sent += 1

                skip_key = f"{brand_key}:{row}:{creator}"
                if (
                    creator
                    and profile_url
                    and not sent_value
                    and not is_excluded_topic
                    and skip_key not in skip_log
                ):
                    pending.append(
                        {
                            "row": row,
                            "instagram_id": creator,
                            "topic": topic,
                            "profile_url": profile_url,
                            "message": "",
                        }
                    )

            message_template = get_dm_message_template(brand_key, brand)
            for target in pending:
                target["message"] = format_message(
                    message_template, target["instagram_id"]
                )

            return {
                "brand_key": brand_key,
                "brand_name": brand["name"],
                "eyebrow": brand["eyebrow"],
                "theme": brand["theme"],
                "expected_account": brand["expected_account"],
                "login_note": brand["login_note"],
                "message_ready": message_template != "__MESSAGE_REQUIRED__",
                "editable_message_template": message_template,
                "sheet": brand["sheet_name"],
                "workbook": CONFIG["_workbook_path"].name,
                "goal": brand["weekly_goal"],
                "weekly_sent": weekly_sent,
                "remaining_to_goal": max(brand["weekly_goal"] - weekly_sent, 0),
                "pending_count": len(pending),
                "excluded_topic_count": excluded_topic_count,
                "excluded_topics": sorted(excluded_topics),
                "targets": pending,
                "dm_sync": sync_status,
                "session_sent": SESSION[brand_key]["sent"],
                "session_skipped": SESSION[brand_key]["skipped"],
                "week": f"{week_start:%Y.%m.%d} ~ {week_end:%Y.%m.%d}",
                "recording": {
                    "sent_date_column": brand["columns"]["sent_date"],
                    "summary_range": brand["summary_label"],
                },
                "reference_image_url": (
                    f"/api/dm/reference-image?brand={brand_key}"
                    if DM_REFERENCE_IMAGES.get(brand_key, Path()).is_file()
                    else ""
                ),
                "reference_image_name": (
                    DM_REFERENCE_IMAGES[brand_key].name
                    if brand_key in DM_REFERENCE_IMAGES
                    and DM_REFERENCE_IMAGES[brand_key].is_file()
                    else ""
                ),
            }
        finally:
            workbook.close()


def brand_connecting_cell_value(value: object) -> object:
    if isinstance(value, datetime):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, date):
        return value.strftime("%Y.%m.%d")
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value if value is not None else ""


def get_brand_connecting(brand_key: str, force_sync: bool = False) -> dict:
    source = BRAND_CONNECTING_SHEETS.get(brand_key)
    if source is None:
        raise ValueError("지원하지 않는 브랜드입니다.")

    sync_status = sync_dm_workbook(force=force_sync)
    workbook_path = CONFIG["_workbook_path"]
    if not workbook_path.is_file():
        raise FileNotFoundError("쇼핑커넥트 원본 파일을 찾을 수 없습니다.")

    with LOCK:
        workbook = load_workbook(
            workbook_path,
            read_only=True,
            data_only=True,
            keep_links=True,
        )
        try:
            sheet_name = source["sheet_name"]
            if sheet_name not in workbook.sheetnames:
                raise KeyError(f"'{sheet_name}' 시트를 찾을 수 없습니다.")
            sheet = workbook[sheet_name]
            sheet.calculate_dimension(force=True)
            column_count = int(source["column_count"])
            sheet_rows = sheet.iter_rows(
                min_row=1,
                max_row=sheet.max_row,
                min_col=1,
                max_col=column_count,
                values_only=True,
            )
            header_values = next(sheet_rows, ())
            headers = [str(value or "").strip() for value in header_values]
            rows = []
            for row_number, raw_values in enumerate(sheet_rows, start=2):
                values = [
                    brand_connecting_cell_value(value)
                    for value in raw_values
                ]
                if not str(values[1] or "").strip():
                    continue
                rows.append({"row": row_number, "values": values})
            return {
                "brand_key": brand_key,
                "brand_name": source["brand_name"],
                "sheet_name": sheet_name,
                "headers": headers,
                "rows": rows,
                "row_count": len(rows),
                "source_url": (
                    "https://docs.google.com/spreadsheets/d/"
                    "1jutw9GQkXuVZ2rlaTX_ZecaF0Mkkalrm/edit"
                ),
                "sync": sync_status,
            }
        finally:
            workbook.close()


def ensure_backup() -> None:
    global BACKUP_CREATED
    if BACKUP_CREATED:
        return
    source = CONFIG["_workbook_path"]
    backup_dir = APP_DIR / "backups"
    backup_dir.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    shutil.copy2(source, backup_dir / f"{source.stem}_backup_{stamp}{source.suffix}")
    BACKUP_CREATED = True


def save_atomic(workbook) -> None:
    destination = CONFIG["_workbook_path"]
    temporary = destination.with_name(f".{destination.stem}.saving{destination.suffix}")
    try:
        workbook.save(temporary)
        os.replace(temporary, destination)
    finally:
        if temporary.exists():
            temporary.unlink()


def upload_dm_workbook_to_drive() -> dict:
    sync_config = dict(CONFIG.get("dm_sync") or {})
    drive_file_id = str(sync_config.get("drive_file_id", "")).strip()
    if not drive_file_id:
        return {"status": "disabled", "message": "원본 Drive 파일 ID가 없습니다."}
    if not DM_DRIVE_TOKEN_PATH.exists():
        return {
            "status": "login_required",
            "message": "Google Drive 원본 연결이 필요합니다.",
        }
    try:
        from googleapiclient.http import MediaFileUpload

        service = load_dm_drive_service()
        media = MediaFileUpload(
            str(CONFIG["_workbook_path"]),
            mimetype=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            resumable=True,
        )
        result = (
            service.files()
            .update(
                fileId=drive_file_id,
                media_body=media,
                fields="id,modifiedTime,size",
            )
            .execute()
        )
        return {
            "status": "completed",
            "message": "Google Drive 원본까지 최신화했습니다.",
            "modified_time": result.get("modifiedTime"),
            "size": result.get("size"),
        }
    except Exception as exc:
        return {
            "status": "error",
            "message": f"로컬 기록은 저장했지만 Drive 최신화에 실패했습니다: {exc}",
        }


def prepare_brand_connect_crawl(brand_key: str) -> set[str]:
    source = BRAND_CONNECTING_SHEETS.get(brand_key)
    if source is None:
        raise ValueError("지원하지 않는 브랜드입니다.")
    sync_dm_workbook(force=True)
    with LOCK:
        workbook = load_workbook(
            CONFIG["_workbook_path"],
            read_only=True,
            data_only=True,
            keep_links=True,
        )
        try:
            sheet = workbook[source["sheet_name"]]
            sheet.calculate_dimension(force=True)
            creators = set()
            for row in sheet.iter_rows(
                min_row=2,
                max_row=sheet.max_row,
                min_col=2,
                max_col=2,
                values_only=True,
            ):
                key = normalize_creator(row[0] if row else "")
                if key:
                    creators.add(key)
            return creators
        finally:
            workbook.close()


def save_brand_connect_results(
    brand_key: str,
    candidates: list[dict],
) -> dict:
    source = BRAND_CONNECTING_SHEETS.get(brand_key)
    if source is None:
        raise ValueError("지원하지 않는 브랜드입니다.")
    sheet_name = source["sheet_name"]
    added = 0
    duplicates = 0
    with LOCK:
        workbook = load_workbook(
            CONFIG["_workbook_path"],
            data_only=False,
            keep_links=True,
        )
        try:
            sheet = workbook[sheet_name]
            existing = {
                normalize_creator(sheet.cell(row, 2).value)
                for row in range(2, sheet.max_row + 1)
                if normalize_creator(sheet.cell(row, 2).value)
            }
            available_rows = [
                row
                for row in range(2, sheet.max_row + 1)
                if not str(sheet.cell(row, 2).value or "").strip()
            ]
            next_append_row = sheet.max_row + 1
            for candidate in candidates:
                creator = str(candidate.get("creator", "")).strip()
                creator_key = normalize_creator(creator)
                if not creator_key or creator_key in existing:
                    duplicates += 1
                    continue
                if available_rows:
                    row = available_rows.pop(0)
                else:
                    row = next_append_row
                    next_append_row += 1
                    template_row = max(2, row - 1)
                    for column in range(2, 6):
                        copy_cell_style(
                            sheet.cell(template_row, column),
                            sheet.cell(row, column),
                        )
                history_count = int(candidate.get("history_count", 0) or 0)
                audience_count = int(candidate.get("audience_count", 0) or 0)
                if brand_key == "alp":
                    values = [
                        creator,
                        str(candidate.get("topics", "")).strip(),
                        history_count,
                        audience_count if audience_count else "",
                        "대기",
                    ]
                else:
                    values = [
                        creator,
                        str(candidate.get("topics", "")).strip(),
                        str(candidate.get("platform", "")).strip(),
                        history_count,
                        audience_count if audience_count else "",
                    ]
                for offset, value in enumerate(values, start=2):
                    sheet.cell(row, offset).value = value
                existing.add(creator_key)
                added += 1
            if added:
                ensure_backup()
                save_atomic(workbook)
        finally:
            workbook.close()
    drive = (
        upload_dm_workbook_to_drive()
        if added
        else {"status": "skipped", "message": "추가된 새 명단이 없습니다."}
    )
    return {
        "sheet_name": sheet_name,
        "added": added,
        "duplicates": duplicates,
        "drive": drive,
    }


def prepare_brand_connect_favorites(
    brand_key: str,
    platform: str,
    product: str,
    count: int,
) -> list[dict]:
    source = BRAND_CONNECTING_SHEETS.get(brand_key)
    product_label = FAVORITE_PRODUCT_LABELS.get(brand_key, {}).get(product)
    platform_label = FAVORITE_PLATFORM_LABELS.get(platform)
    if source is None or not product_label or not platform_label:
        raise ValueError("올바른 브랜드·채널·상품을 선택해주세요.")
    sync_dm_workbook(force=True)
    with LOCK:
        workbook = load_workbook(
            CONFIG["_workbook_path"],
            read_only=True,
            data_only=True,
            keep_links=True,
        )
        try:
            sheet = workbook[source["sheet_name"]]
            sheet.calculate_dimension(force=True)
            headers = {}
            for column in range(
                1,
                max(int(source["column_count"]), int(sheet.max_column or 0)) + 1,
            ):
                header = str(sheet.cell(1, column).value or "").strip()
                if header:
                    headers.setdefault(header, column)
            creator_column = headers.get("크리에이터")
            date_column = find_favorite_date_column(
                headers,
                brand_key,
                product,
            )
            platform_column = headers.get("플랫폼")
            product_column = headers.get(product_label)
            if (
                not creator_column
                or not date_column
                or not platform_column
                or not product_column
            ):
                raise KeyError(
                    f"'{source['sheet_name']}' 시트에서 선택 상품의 제안날짜·크리에이터·"
                    f"플랫폼·{product_label} 열을 찾지 못했습니다."
                )
            candidates = []
            for row in range(2, sheet.max_row + 1):
                creator = str(sheet.cell(row, creator_column).value or "").strip()
                row_platform = str(
                    sheet.cell(row, platform_column).value or ""
                ).strip()
                proposal_date = sheet.cell(row, date_column).value
                selected = str(sheet.cell(row, product_column).value or "").strip()
                if (
                    not creator
                    or row_platform != platform_label
                    or not is_favorite_candidate(proposal_date, selected)
                ):
                    continue
                candidates.append({"row": row, "creator": creator})
                if len(candidates) >= count:
                    break
            return candidates
        finally:
            workbook.close()


def save_brand_connect_favorites(
    brand_key: str,
    product: str,
    completed: list[dict],
) -> dict:
    source = BRAND_CONNECTING_SHEETS.get(brand_key)
    product_label = FAVORITE_PRODUCT_LABELS.get(brand_key, {}).get(product)
    if source is None or not product_label:
        raise ValueError("올바른 브랜드·상품을 선택해주세요.")
    marked = 0
    dated = 0
    date_preserved = 0
    skipped = 0
    favorite_date = date.today().strftime("%Y.%m.%d")
    with LOCK:
        workbook = load_workbook(
            CONFIG["_workbook_path"],
            data_only=False,
            keep_links=True,
        )
        try:
            sheet = workbook[source["sheet_name"]]
            headers = {}
            for column in range(1, sheet.max_column + 1):
                header = str(sheet.cell(1, column).value or "").strip()
                if header:
                    headers.setdefault(header, column)
            creator_column = headers.get("크리에이터")
            date_column = find_favorite_date_column(
                headers,
                brand_key,
                product,
            )
            product_column = headers.get(product_label)
            if not creator_column or not date_column or not product_column:
                raise KeyError(
                    f"'{source['sheet_name']}' 시트에서 선택 상품의 제안날짜·크리에이터·"
                    f"{product_label} 열을 찾지 못했습니다."
                )
            for item in completed:
                row = int(item.get("row", 0) or 0)
                creator = str(item.get("creator", "")).strip()
                if row < 2 or normalize_creator(
                    sheet.cell(row, creator_column).value
                ) != normalize_creator(creator):
                    skipped += 1
                    continue
                cell = sheet.cell(row, product_column)
                if str(cell.value or "").strip():
                    skipped += 1
                    continue
                cell.value = "찜"
                date_preserved += int(
                    bool(str(sheet.cell(row, int(date_column)).value or "").strip())
                )
                marked += 1
            if marked:
                ensure_backup()
                save_atomic(workbook)
        finally:
            workbook.close()
    drive = (
        upload_dm_workbook_to_drive()
        if marked
        else {"status": "skipped", "message": "새로 표시할 명단이 없습니다."}
    )
    return {
        "sheet_name": source["sheet_name"],
        "marked": marked,
        "dated": dated,
        "date_preserved": date_preserved,
        "favorite_date": favorite_date,
        "skipped": skipped,
        "drive": drive,
    }


def rebuild_brand_connect_proposal_summary(
    sheet,
    brand_key: str,
    headers: dict[str, int],
    product_columns: dict[str, int | None],
) -> tuple[int, bool]:
    products = tuple(product_columns)
    product_labels = PROPOSAL_PRODUCT_LABELS[brand_key]
    date_columns = {
        product: find_favorite_date_column(headers, brand_key, product)
        for product in products
    }
    summary_date_column = next(
        (
            column
            for column in range(1, sheet.max_column + 1)
            if str(sheet.cell(1, column).value or "").strip() == "날짜"
        ),
        None,
    )
    if not summary_date_column:
        return 0, False
    summary_columns: dict[str, int] = {}
    for column in range(summary_date_column + 1, sheet.max_column + 1):
        compact = re.sub(
            r"\s+",
            "",
            str(sheet.cell(1, column).value or ""),
        )
        for product, label in product_labels.items():
            if compact == re.sub(r"\s+", "", label):
                summary_columns[product] = column
    if (
        set(summary_columns) != set(products)
        or not all(date_columns.values())
    ):
        return 0, False

    counts: dict[str, dict[str, int]] = {}
    creator_column = headers.get("크리에이터")
    if not creator_column:
        return 0, False
    for row in range(2, sheet.max_row + 1):
        if not str(sheet.cell(row, creator_column).value or "").strip():
            continue
        for product in products:
            status = str(
                sheet.cell(row, int(product_columns[product])).value or ""
            ).strip()
            if status not in PROPOSAL_STATUS_VALUES:
                continue
            raw_date = sheet.cell(row, int(date_columns[product])).value
            try:
                proposal_date = normalize_proposal_date(raw_date)
            except ValueError:
                continue
            counts.setdefault(
                proposal_date,
                {item: 0 for item in products},
            )[product] += 1

    dates = sorted(
        counts,
        key=lambda value: datetime.strptime(value, "%Y.%m.%d"),
    )
    last_existing = 1
    for row in range(2, min(sheet.max_row, 200) + 1):
        if any(
            sheet.cell(row, column).value not in (None, "")
            for column in (
                summary_date_column,
                *(summary_columns[product] for product in products),
            )
        ):
            last_existing = row
    last_target = max(last_existing, len(dates) + 1)
    changed = False
    style_row = 2
    for row in range(2, last_target + 1):
        values = (
            (
                dates[row - 2],
                *(counts[dates[row - 2]][product] for product in products),
            )
            if row - 2 < len(dates)
            else tuple(None for _ in range(len(products) + 1))
        )
        for column, value in zip(
            (
                summary_date_column,
                *(summary_columns[product] for product in products),
            ),
            values,
        ):
            cell = sheet.cell(row, column)
            if cell.value != value:
                cell.value = value
                changed = True
            if row > style_row:
                copy_cell_style(sheet.cell(style_row, column), cell)
    return len(dates), changed


def save_brand_connect_proposals(
    brand_key: str,
    results: list[dict],
    reconciliation_campaigns: list[dict] | None = None,
) -> dict:
    source = BRAND_CONNECTING_SHEETS.get(brand_key)
    product_labels = PROPOSAL_PRODUCT_LABELS.get(brand_key, {})
    if source is None or not product_labels:
        raise ValueError("지원하지 않는 브랜드입니다.")

    def relaxed_creator(value: object) -> str:
        return re.sub(r"[^0-9a-z가-힣]", "", str(value or "").casefold())

    matched = 0
    updated = 0
    links = 0
    dated = 0
    reconciled = 0
    summary_rows = 0
    unmatched_names: list[str] = []
    changed = False
    with LOCK:
        workbook = load_workbook(
            CONFIG["_workbook_path"],
            data_only=False,
            keep_links=True,
        )
        try:
            sheet = workbook[source["sheet_name"]]
            headers = {}
            for column in range(1, sheet.max_column + 1):
                header = str(sheet.cell(1, column).value or "").strip()
                if header:
                    headers.setdefault(header, column)
            creator_column = headers.get("크리에이터")
            content_header = "콘텐츠 제출" if brand_key == "alp" else "콘텐츠 링크"
            content_column = headers.get(content_header)
            product_columns = {
                product: headers.get(label)
                for product, label in product_labels.items()
            }
            date_columns = {
                product: find_favorite_date_column(headers, brand_key, product)
                for product in product_labels
            }
            if not creator_column:
                raise KeyError(
                    f"'{source['sheet_name']}' 시트에서 크리에이터 열을 찾지 못했습니다."
                )
            missing_products = [
                label
                for product, label in product_labels.items()
                if not product_columns.get(product)
            ]
            if missing_products:
                raise KeyError(
                    f"'{source['sheet_name']}' 시트에서 "
                    f"{', '.join(missing_products)} 열을 찾지 못했습니다."
                )
            missing_dates = [
                product_labels[product]
                for product, column in date_columns.items()
                if not column
            ]
            if missing_dates:
                raise KeyError(
                    f"'{source['sheet_name']}' 시트에서 "
                    f"{', '.join(missing_dates)} 제안날짜 열을 찾지 못했습니다."
                )
            if not content_column:
                raise KeyError(
                    f"'{source['sheet_name']}' 시트에서 {content_header} 열을 찾지 못했습니다."
                )

            exact_rows: dict[str, list[int]] = {}
            relaxed_rows: dict[str, list[int]] = {}
            for row in range(2, sheet.max_row + 1):
                creator = str(sheet.cell(row, creator_column).value or "").strip()
                if not creator:
                    continue
                exact_rows.setdefault(normalize_creator(creator), []).append(row)
                relaxed = relaxed_creator(creator)
                if relaxed:
                    relaxed_rows.setdefault(relaxed, []).append(row)

            merged: dict[tuple[str, str], dict] = {}
            for item in results:
                creator = str(item.get("creator", "")).strip()
                product = str(item.get("product", "")).strip()
                if not creator or product not in product_columns:
                    continue
                merged[(product, normalize_creator(creator))] = dict(item)

            verified_rows: dict[str, set[int]] = {
                product: set() for product in product_labels
            }
            for (product, _), item in merged.items():
                creator = str(item.get("creator", "")).strip()
                rows = exact_rows.get(normalize_creator(creator), [])
                if not rows:
                    relaxed = relaxed_creator(creator)
                    relaxed_matches = relaxed_rows.get(relaxed, []) if relaxed else []
                    if len(relaxed_matches) == 1:
                        rows = relaxed_matches
                if not rows:
                    unmatched_names.append(creator)
                    continue
                status = str(item.get("status", "")).strip()
                content_url = str(item.get("content_url", "")).strip()
                raw_proposal_date = str(
                    item.get("proposal_date", "")
                ).strip()
                proposal_date = (
                    normalize_proposal_date(raw_proposal_date)
                    if raw_proposal_date
                    else ""
                )
                for row in rows:
                    matched += 1
                    verified_rows[product].add(row)
                    status_cell = sheet.cell(row, int(product_columns[product]))
                    if status and str(status_cell.value or "").strip() != status:
                        status_cell.value = status
                        updated += 1
                        changed = True
                    date_cell = sheet.cell(row, int(date_columns[product]))
                    if proposal_date and str(
                        date_cell.value or ""
                    ).strip() != proposal_date:
                        date_cell.value = proposal_date
                        dated += 1
                        changed = True
                    if content_url:
                        content_cell = sheet.cell(row, content_column)
                        if str(content_cell.value or "").strip() != content_url:
                            content_cell.value = content_url
                            content_cell.hyperlink = content_url
                            content_cell.style = "Hyperlink"
                            links += 1
                            changed = True
            reconcile_products = {
                str(campaign.get("product", "")).strip()
                for campaign in (reconciliation_campaigns or [])
                if str(campaign.get("product", "")).strip() in product_columns
                and str(campaign.get("proposal_date", "")).strip()
            }
            for product in reconcile_products:
                status_column = int(product_columns[product])
                date_column = int(date_columns[product])
                for row in range(2, sheet.max_row + 1):
                    if not str(
                        sheet.cell(row, creator_column).value or ""
                    ).strip():
                        continue
                    status_cell = sheet.cell(row, status_column)
                    current_status = str(status_cell.value or "").strip()
                    if row in verified_rows[product]:
                        continue
                    date_cell = sheet.cell(row, date_column)
                    if current_status in PROPOSAL_STATUS_VALUES:
                        status_cell.value = "찜"
                        reconciled += 1
                        changed = True
                    if current_status in PROPOSAL_STATUS_VALUES or current_status == "찜":
                        if str(date_cell.value or "").strip():
                            date_cell.value = None
                            changed = True
            if brand_key in {"alp", "gaia"}:
                summary_rows, summary_changed = rebuild_brand_connect_proposal_summary(
                    sheet,
                    brand_key,
                    headers,
                    product_columns,
                )
                changed = changed or summary_changed
            if changed:
                workbook.calculation.calcMode = "auto"
                workbook.calculation.fullCalcOnLoad = True
                workbook.calculation.forceFullCalc = True
                ensure_backup()
                save_atomic(workbook)
        finally:
            workbook.close()
    drive = (
        upload_dm_workbook_to_drive()
        if changed
        else {"status": "skipped", "message": "새로 반영할 제안 결과가 없습니다."}
    )
    return {
        "sheet_name": source["sheet_name"],
        "matched": matched,
        "updated": updated,
        "links": links,
        "dated": dated,
        "reconciled": reconciled,
        "summary_rows": summary_rows,
        "unmatched": len(unmatched_names),
        "unmatched_names": unmatched_names[:30],
        "drive": drive,
    }


def backfill_brand_connect_proposal_dates(
    proposal_date: str | None = None,
    sync_first: bool = True,
) -> dict:
    return {
        "date": proposal_date or "",
        "alp": 0,
        "gaia": 0,
        "total": 0,
        "proposal_dates_untouched": True,
        "drive": {
            "status": "skipped",
            "message": "제안날짜 자동 입력이 비활성화되어 있습니다.",
        },
    }


BRAND_CONNECT_MANAGER = BrandConnectCrawlerManager(
    APP_DIR,
    prepare_brand_connect_crawl,
    save_brand_connect_results,
)
BRAND_CONNECT_FAVORITE_MANAGER = BrandConnectFavoriteManager(
    APP_DIR,
    prepare_brand_connect_favorites,
    save_brand_connect_favorites,
)
BRAND_CONNECT_PROPOSAL_MANAGER = BrandConnectProposalManager(
    APP_DIR,
    lambda: sync_dm_workbook(force=True),
    save_brand_connect_proposals,
)
BRAND_CONNECT_CAMPAIGNS_LOCK = threading.Lock()


def load_brand_connect_campaigns(brand: str) -> dict:
    if brand not in {"alp", "gaia"}:
        raise ValueError("지원하지 않는 브랜드입니다.")
    with BRAND_CONNECT_CAMPAIGNS_LOCK:
        try:
            stored = json.loads(
                BRAND_CONNECT_CAMPAIGNS_PATH.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError):
            stored = {}
    rows = stored.get(brand, []) if isinstance(stored, dict) else []
    if not isinstance(rows, list):
        rows = []
    allowed_products = {
        "alp": {"immun", "iron_drop"},
        "gaia": {"oil", "pickles", "pouch"},
    }[brand]
    campaigns = []
    for row in rows[:20]:
        if not isinstance(row, dict):
            continue
        product = str(row.get("product", "")).strip()
        if product not in allowed_products:
            continue
        campaigns.append(
            {
                "url": str(row.get("url", ""))[:1000],
                "product": product,
                "proposal_date": str(row.get("proposal_date", "")).strip(),
            }
        )
    if any(
        not campaign["proposal_date"] for campaign in campaigns
    ):
        sync_dm_workbook(force=True)
        with LOCK:
            workbook = load_workbook(
                CONFIG["_workbook_path"],
                data_only=True,
                keep_links=True,
            )
            try:
                sheet = workbook[BRAND_CONNECTING_SHEETS[brand]["sheet_name"]]
                product_labels = PROPOSAL_PRODUCT_LABELS[brand]
                date_queues = {product: [] for product in product_labels}
                summary_date_column = next(
                    (
                        column
                        for column in range(1, sheet.max_column + 1)
                        if str(sheet.cell(1, column).value or "").strip()
                        == "날짜"
                    ),
                    None,
                )
                summary_columns: dict[str, int] = {}
                if summary_date_column:
                    for column in range(
                        summary_date_column + 1,
                        sheet.max_column + 1,
                    ):
                        compact = re.sub(
                            r"\s+",
                            "",
                            str(sheet.cell(1, column).value or ""),
                        )
                        for product, label in product_labels.items():
                            if compact == re.sub(r"\s+", "", label):
                                summary_columns[product] = column
                for row in range(2, min(sheet.max_row, 200) + 1):
                    if not summary_date_column:
                        break
                    raw_date = sheet.cell(row, summary_date_column).value
                    if raw_date in (None, ""):
                        continue
                    try:
                        proposal_date = normalize_proposal_date(raw_date)
                    except ValueError:
                        continue
                    for product, column in summary_columns.items():
                        if int(sheet.cell(row, column).value or 0) > 0:
                            date_queues[product].append(proposal_date)
                queue_indexes = {product: 0 for product in product_labels}
                for campaign in campaigns:
                    if campaign["proposal_date"]:
                        continue
                    product = campaign["product"]
                    index = queue_indexes[product]
                    if index < len(date_queues[product]):
                        campaign["proposal_date"] = date_queues[product][index]
                        queue_indexes[product] += 1
            finally:
                workbook.close()
    return {"brand": brand, "campaigns": campaigns}


def save_brand_connect_campaigns(payload: dict) -> dict:
    brand = str(payload.get("brand", "")).strip()
    if brand not in {"alp", "gaia"}:
        raise ValueError("지원하지 않는 브랜드입니다.")
    rows = payload.get("campaigns")
    if not isinstance(rows, list):
        raise ValueError("캠페인 목록 형식이 올바르지 않습니다.")
    if len(rows) > 20:
        raise ValueError("캠페인 링크는 최대 20개까지 저장할 수 있습니다.")
    allowed_products = {
        "alp": {"immun", "iron_drop"},
        "gaia": {"oil", "pickles", "pouch"},
    }[brand]
    campaigns = []
    for row in rows:
        if not isinstance(row, dict):
            raise ValueError("캠페인 행 형식이 올바르지 않습니다.")
        product = str(row.get("product", "")).strip()
        if product not in allowed_products:
            raise ValueError("브랜드에 맞는 상품을 선택해주세요.")
        campaigns.append(
            {
                "url": str(row.get("url", "")).strip()[:1000],
                "product": product,
                "proposal_date": (
                    normalize_proposal_date(row.get("proposal_date"))
                    if str(row.get("proposal_date", "")).strip()
                    else ""
                ),
            }
        )
    with BRAND_CONNECT_CAMPAIGNS_LOCK:
        try:
            stored = json.loads(
                BRAND_CONNECT_CAMPAIGNS_PATH.read_text(encoding="utf-8")
            )
        except (OSError, json.JSONDecodeError):
            stored = {}
        if not isinstance(stored, dict):
            stored = {}
        stored[brand] = campaigns
        BRAND_CONNECT_CAMPAIGNS_PATH.parent.mkdir(parents=True, exist_ok=True)
        temporary = BRAND_CONNECT_CAMPAIGNS_PATH.with_suffix(".tmp")
        temporary.write_text(
            json.dumps(stored, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        temporary.replace(BRAND_CONNECT_CAMPAIGNS_PATH)
    return {"ok": True, "brand": brand, "campaigns": campaigns}


def copy_cell_style(source, destination) -> None:
    destination.font = copy(source.font)
    destination.fill = copy(source.fill)
    destination.border = copy(source.border)
    destination.alignment = copy(source.alignment)
    destination.number_format = source.number_format
    destination.protection = copy(source.protection)


def update_proposal_summary(
    sheet, brand: dict, proposed_date: str, columns: dict[str, int]
) -> int:
    summary = brand["proposal_summary"]
    date_column = summary["date_column"]
    count_rules = summary["count_rules"]
    total_row = summary["total_row"]
    first_date_row = summary["first_date_row"]
    last_summary_column = max(rule["summary_column"] for rule in count_rules)

    data_last_row = max(
        row
        for row in range(2, sheet.max_row + 1)
        if sheet.cell(row, columns["creator"]).value not in (None, "")
    )

    last_date_row = first_date_row - 1
    target_summary_row = None
    row = first_date_row
    while sheet.cell(row, date_column).value not in (None, ""):
        date_text = str(sheet.cell(row, date_column).value).strip()
        last_date_row = row
        if date_text == proposed_date:
            target_summary_row = row
        row += 1

    if target_summary_row is None:
        target_summary_row = last_date_row + 1
        style_source_row = last_date_row if last_date_row >= first_date_row else first_date_row
        for column in range(date_column, last_summary_column + 1):
            copy_cell_style(
                sheet.cell(style_source_row, column),
                sheet.cell(target_summary_row, column),
            )
        sheet.cell(target_summary_row, date_column).value = proposed_date
        last_date_row = target_summary_row

    date_letter = get_column_letter(date_column)
    for formula_row in range(first_date_row, last_date_row + 1):
        if sheet.cell(formula_row, date_column).value in (None, ""):
            continue
        for rule in count_rules:
            source_letter = get_column_letter(rule["source_column"])
            sheet.cell(formula_row, rule["summary_column"]).value = (
                f"=COUNTIFS(${source_letter}$2:${source_letter}${data_last_row},"
                f"${date_letter}{formula_row})"
            )

    for rule in count_rules:
        summary_letter = get_column_letter(rule["summary_column"])
        sheet.cell(total_row, rule["summary_column"]).value = (
            f"=SUM({summary_letter}{first_date_row}:"
            f"{summary_letter}{last_date_row})"
        )
    return target_summary_row


def update_target(brand_key: str, row: int, action: str) -> dict:
    brand = get_brand(brand_key)
    with LOCK:
        workbook, sheet = open_sheet(brand)
        try:
            columns = required_columns(sheet, brand)
            creator = normalize_instagram_id(sheet.cell(row, columns["creator"]).value)
            if not creator:
                raise ValueError(f"{row}행의 크리에이터 ID가 비어 있습니다.")

            if action == "sent":
                ensure_backup()
                sent_date = date.today().strftime("%Y.%m.%d")
                sheet.cell(row, columns["sent_date"]).value = sent_date
                summary_row = update_proposal_summary(
                    sheet, brand, sent_date, columns
                )
                workbook.calculation.calcMode = "auto"
                workbook.calculation.fullCalcOnLoad = True
                workbook.calculation.forceFullCalc = True
                SESSION[brand_key]["sent"] += 1
                save_atomic(workbook)
                drive_sync = upload_dm_workbook_to_drive()
            elif action == "skip":
                skip_log = load_skip_log()
                skip_log[f"{brand_key}:{row}:{creator}"] = {
                    "brand": brand_key,
                    "row": row,
                    "instagram_id": creator,
                    "skipped_at": datetime.now().isoformat(timespec="seconds"),
                }
                save_skip_log(skip_log)
                SESSION[brand_key]["skipped"] += 1
            else:
                raise ValueError("지원하지 않는 작업입니다.")

            result = {
                "ok": True,
                "brand": brand_key,
                "row": row,
                "action": action,
                "instagram_id": creator,
            }
            if action == "sent":
                result["summary_row"] = summary_row
                result["drive_sync"] = drive_sync
            return result
        except PermissionError as exc:
            raise PermissionError(
                "Excel 파일이 열려 있어 저장할 수 없습니다. Excel에서 작업용 파일을 닫고 다시 눌러주세요."
            ) from exc
        finally:
            workbook.close()


INSTAGRAM_DM_SENDER = InstagramDMSenderManager(
    APP_DIR,
    get_dashboard,
    update_target,
    CONFIG["brands"],
    DM_REFERENCE_IMAGES,
)


class RequestHandler(BaseHTTPRequestHandler):
    def log_message(self, format_string: str, *args) -> None:
        return

    def send_json(self, payload: dict, status: int = 200) -> None:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_html(self, path: Path) -> None:
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "text/html; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_css(self, path: Path) -> None:
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "text/css; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_file(self, path: Path, filename: str) -> None:
        body = path.read_bytes()
        self.send_response(200)
        self.send_header(
            "Content-Type",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.send_header(
            "Content-Disposition",
            f"attachment; filename*=UTF-8''{quote(filename)}",
        )
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_png(self, path: Path) -> None:
        if not path.is_file():
            self.send_json({"error": "공동구매 구성안 이미지를 찾을 수 없습니다."}, 404)
            return
        body = path.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", "image/png")
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def send_audio(self, path: Path, mime_type: str) -> None:
        if not path.is_file():
            self.send_json({"error": "녹음 파일을 찾을 수 없습니다."}, 404)
            return
        size = path.stat().st_size
        start = 0
        end = size - 1
        status = 200
        range_header = self.headers.get("Range", "")
        if range_header.startswith("bytes="):
            requested = range_header[6:].split(",", 1)[0]
            start_text, _, end_text = requested.partition("-")
            try:
                if start_text:
                    start = int(start_text)
                if end_text:
                    end = min(int(end_text), size - 1)
                if start < 0 or start > end or start >= size:
                    raise ValueError
                status = 206
            except ValueError:
                self.send_response(416)
                self.send_header("Content-Range", f"bytes */{size}")
                self.end_headers()
                return
        length = end - start + 1
        self.send_response(status)
        self.send_header("Content-Type", mime_type or "audio/webm")
        self.send_header("Content-Length", str(length))
        self.send_header("Accept-Ranges", "bytes")
        self.send_header("Cache-Control", "private, max-age=3600")
        if status == 206:
            self.send_header("Content-Range", f"bytes {start}-{end}/{size}")
        self.end_headers()
        with path.open("rb") as audio:
            audio.seek(start)
            remaining = length
            while remaining:
                chunk = audio.read(min(1024 * 1024, remaining))
                if not chunk:
                    break
                self.wfile.write(chunk)
                remaining -= len(chunk)

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            if path == "/":
                self.send_html(HTML_PATH)
            elif path == "/prices":
                self.send_html(PRICE_HTML_PATH)
            elif path == "/simulation":
                self.send_html(SIMULATION_HTML_PATH)
            elif path == "/calendar":
                self.send_html(CALENDAR_HTML_PATH)
            elif path == "/meetings":
                self.send_html(MEETINGS_HTML_PATH)
            elif path == "/memos":
                self.send_html(MEMOS_HTML_PATH)
            elif path == "/sales-email":
                self.send_html(SALES_EMAIL_HTML_PATH)
            elif path == "/theme-meeting.css":
                self.send_css(THEME_CSS_PATH)
            elif path == "/brand-connecting":
                self.send_html(BRAND_CONNECTING_HTML_PATH)
            elif path == "/api/dashboard":
                brand_key = parse_qs(parsed.query).get("brand", ["alp"])[0]
                force_sync = parse_qs(parsed.query).get("force_sync", ["0"])[0] == "1"
                self.send_json(get_dashboard(brand_key, force_sync=force_sync))
            elif path == "/api/dm/reference-image":
                brand_key = parse_qs(parsed.query).get("brand", [""])[0]
                image_path = DM_REFERENCE_IMAGES.get(brand_key)
                if image_path is None:
                    self.send_json({"error": "지원하지 않는 브랜드입니다."}, 404)
                else:
                    self.send_png(image_path)
            elif path == "/api/health":
                self.send_json({"ok": True, "brands": list(CONFIG["brands"])})
            elif path == "/api/prices/status":
                self.send_json(PRICE_MANAGER.snapshot())
            elif path == "/api/dm/auto/status":
                self.send_json(INSTAGRAM_DM_SENDER.snapshot())
            elif path == "/api/calendar":
                month = parse_qs(parsed.query).get(
                    "month", [date.today().strftime("%Y-%m")]
                )[0]
                self.send_json(get_calendar_month(month))
            elif path == "/api/calendar/today":
                self.send_json(get_calendar_today())
            elif path == "/api/calendar/routines":
                self.send_json(get_calendar_checklist_routines())
            elif path == "/api/calendar/sync/status":
                self.send_json(dict(CALENDAR_SYNC_STATUS))
            elif path == "/api/meetings":
                self.send_json(get_meeting_notes())
            elif path == "/api/memos":
                self.send_json(get_platform_memos())
            elif path == "/api/sales-email":
                self.send_json(public_sales_email_data())
            elif path == "/api/brand-connecting":
                brand_key = parse_qs(parsed.query).get("brand", ["alp"])[0]
                force_sync = (
                    parse_qs(parsed.query).get("force_sync", ["0"])[0] == "1"
                )
                self.send_json(
                    get_brand_connecting(brand_key, force_sync=force_sync)
                )
            elif path == "/api/brand-connecting/crawl/status":
                self.send_json(BRAND_CONNECT_MANAGER.snapshot())
            elif path == "/api/brand-connecting/favorite/status":
                self.send_json(BRAND_CONNECT_FAVORITE_MANAGER.snapshot())
            elif path == "/api/brand-connecting/proposal/status":
                self.send_json(BRAND_CONNECT_PROPOSAL_MANAGER.snapshot())
            elif path == "/api/brand-connecting/proposal/campaigns":
                brand_key = parse_qs(parsed.query).get("brand", ["alp"])[0]
                self.send_json(load_brand_connect_campaigns(brand_key))
            elif path == "/api/meetings/ai/status":
                self.send_json(meeting_ai_status())
            elif path == "/api/meetings/ai/job":
                note_id = str(parse_qs(parsed.query).get("note_id", [""])[0])
                self.send_json(public_meeting_ai_job(note_id))
            elif path.startswith("/api/meetings/recordings/"):
                parts = path.strip("/").split("/")
                if len(parts) != 5:
                    self.send_json({"error": "찾을 수 없습니다."}, 404)
                    return
                note_id, recording_id = parts[3], parts[4]
                with MEETING_NOTES_LOCK:
                    notes = load_meeting_notes()
                    note = find_meeting_note(notes, note_id)
                    recording = next(
                        (
                            item
                            for item in note.get("recordings", [])
                            if item.get("id") == recording_id
                        ),
                        None,
                    )
                if recording is None:
                    self.send_json({"error": "녹음을 찾을 수 없습니다."}, 404)
                    return
                recording_path = (
                    MEETING_RECORDINGS_DIR
                    / note_id
                    / str(recording.get("filename", ""))
                )
                self.send_audio(
                    recording_path,
                    str(recording.get("mime_type", "audio/webm")),
                )
            else:
                self.send_json({"error": "찾을 수 없습니다."}, 404)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 500)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        try:
            length = int(self.headers.get("Content-Length", "0"))
            if path == "/api/meetings/recordings":
                query = parse_qs(parsed.query)
                note_id = str(query.get("note_id", [""])[0]).strip()
                mime_type = self.headers.get(
                    "Content-Type",
                    str(query.get("mime_type", ["audio/webm"])[0]),
                )
                duration = int(query.get("duration", ["0"])[0])
                source = str(query.get("source", ["meeting"])[0])
                self.send_json(
                    save_meeting_recording(
                        note_id,
                        mime_type,
                        duration,
                        self.rfile,
                        length,
                        source,
                    )
                )
                return
            if path == "/api/sales-email/attachments":
                query = parse_qs(parsed.query)
                filename = str(query.get("filename", [""])[0])
                mime_type = self.headers.get(
                    "Content-Type",
                    str(query.get("mime_type", ["application/octet-stream"])[0]),
                )
                self.send_json(
                    save_sales_email_attachment(
                        filename,
                        mime_type,
                        self.rfile,
                        length,
                    )
                )
                return
            payload = json.loads(self.rfile.read(length) or b"{}")
            if path == "/api/target":
                brand_key = str(payload["brand"])
                row = int(payload["row"])
                action = str(payload["action"])
                self.send_json(update_target(brand_key, row, action))
            elif path == "/api/dm/template":
                brand_key = str(payload.get("brand", "")).strip()
                action = str(payload.get("action", "save")).strip()
                if action == "reset":
                    self.send_json(reset_dm_message_template(brand_key))
                else:
                    self.send_json(
                        save_dm_message_template(
                            brand_key,
                            str(payload.get("template", "")),
                        )
                    )
            elif path == "/api/dm/sync-drive":
                self.send_json(upload_dm_workbook_to_drive())
            elif path == "/api/dm/auto/start":
                self.send_json(
                    INSTAGRAM_DM_SENDER.start(
                        str(payload.get("brand", "")).strip()
                    )
                )
            elif path == "/api/dm/auto/resume":
                self.send_json(INSTAGRAM_DM_SENDER.resume_after_login())
            elif path == "/api/dm/auto/stop":
                self.send_json(INSTAGRAM_DM_SENDER.stop())
            elif path == "/api/brand-connecting/crawl/start":
                if (
                    BRAND_CONNECT_FAVORITE_MANAGER.is_running()
                    or BRAND_CONNECT_PROPOSAL_MANAGER.is_running()
                ):
                    raise ValueError(
                        "다른 브랜드커넥트 자동화가 진행 중입니다. 완료 후 명단 수집을 시작해주세요."
                    )
                source = str(payload.get("source", "both"))
                platforms = {
                    "blog": ["blog"],
                    "instagram": ["instagram"],
                    "both": ["blog", "instagram"],
                }.get(source)
                if platforms is None:
                    raise ValueError("올바른 크롤링 대상을 선택해주세요.")
                self.send_json(
                    BRAND_CONNECT_MANAGER.start(
                        str(payload.get("brand", "")),
                        platforms,
                        int(payload.get("count", 100)),
                    )
                )
            elif path == "/api/brand-connecting/crawl/resume":
                self.send_json(BRAND_CONNECT_MANAGER.resume_after_login())
            elif path == "/api/brand-connecting/crawl/stop":
                self.send_json(BRAND_CONNECT_MANAGER.stop())
            elif path == "/api/brand-connecting/favorite/start":
                if BRAND_CONNECT_MANAGER.snapshot().get(
                    "status"
                ) in FAVORITE_RUNNING_STATUSES or BRAND_CONNECT_PROPOSAL_MANAGER.is_running():
                    raise ValueError(
                        "다른 브랜드커넥트 자동화가 진행 중입니다. 완료 후 찜 자동화를 시작해주세요."
                    )
                self.send_json(
                    BRAND_CONNECT_FAVORITE_MANAGER.start(
                        str(payload.get("brand", "")),
                        str(payload.get("platform", "")),
                        str(payload.get("product", "")),
                        int(payload.get("count", 10)),
                    )
                )
            elif path == "/api/brand-connecting/favorite/resume":
                self.send_json(
                    BRAND_CONNECT_FAVORITE_MANAGER.resume_after_login()
                )
            elif path == "/api/brand-connecting/favorite/stop":
                self.send_json(BRAND_CONNECT_FAVORITE_MANAGER.stop())
            elif path == "/api/brand-connecting/proposal/start":
                if (
                    BRAND_CONNECT_MANAGER.snapshot().get("status")
                    in PROPOSAL_RUNNING_STATUSES
                    or BRAND_CONNECT_FAVORITE_MANAGER.is_running()
                ):
                    raise ValueError(
                        "다른 브랜드커넥트 자동화가 진행 중입니다. 완료 후 제안 확인을 시작해주세요."
                    )
                campaigns = payload.get("campaigns")
                if not isinstance(campaigns, list):
                    raise ValueError("확인할 캠페인 링크를 입력해주세요.")
                self.send_json(
                    BRAND_CONNECT_PROPOSAL_MANAGER.start(
                        str(payload.get("brand", "")),
                        campaigns,
                    )
                )
            elif path == "/api/brand-connecting/proposal/resume":
                self.send_json(
                    BRAND_CONNECT_PROPOSAL_MANAGER.resume_after_login()
                )
            elif path == "/api/brand-connecting/proposal/stop":
                self.send_json(BRAND_CONNECT_PROPOSAL_MANAGER.stop())
            elif path == "/api/brand-connecting/proposal/dates/backfill":
                self.send_json(
                    backfill_brand_connect_proposal_dates(
                        str(payload.get("date", "")).strip() or None,
                        bool(payload.get("sync_first", True)),
                    )
                )
            elif path == "/api/brand-connecting/proposal/campaigns":
                self.send_json(save_brand_connect_campaigns(payload))
            elif path == "/api/prices/start":
                self.send_json(PRICE_MANAGER.start())
            elif path == "/api/prices/resume":
                self.send_json(PRICE_MANAGER.resume_after_login())
            elif path == "/api/prices/stop":
                self.send_json(PRICE_MANAGER.stop())
            elif path == "/api/simulation/generate":
                output_path, filename = SIMULATION_MANAGER.generate(payload)
                self.send_file(output_path, filename)
            elif path == "/api/calendar/task":
                self.send_json(update_calendar_task(payload))
            elif path == "/api/calendar/routine":
                self.send_json(update_calendar_checklist_routine(payload))
            elif path == "/api/calendar/sync":
                self.send_json(sync_calendar_events())
            elif path == "/api/meetings":
                self.send_json(update_meeting_note(payload))
            elif path == "/api/memos":
                self.send_json(update_platform_memo(payload))
            elif path == "/api/sales-email":
                self.send_json(update_sales_email_data(payload))
            elif path == "/api/sales-email/outlook-draft":
                self.send_json(create_outlook_draft(payload))
            elif path == "/api/meetings/recordings/delete":
                self.send_json(
                    delete_meeting_recording(
                        str(payload.get("note_id", "")),
                        str(payload.get("recording_id", "")),
                    )
                )
            elif path == "/api/meetings/ai/generate":
                self.send_json(
                    start_meeting_ai_job(str(payload.get("note_id", "")))
                )
            else:
                self.send_json({"error": "찾을 수 없습니다."}, 404)
        except Exception as exc:
            self.send_json({"error": str(exc)}, 400)


def main() -> None:
    server = ThreadingHTTPServer((HOST, PORT), RequestHandler)
    url = f"http://{HOST}:{PORT}"
    print("Instagram DM 도우미가 실행되었습니다.")
    print(f"브라우저 주소: {url}")
    print("종료하려면 이 창에서 Ctrl+C를 누르세요.")
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n종료합니다.")
    finally:
        server.server_close()


def run_server_silently() -> None:
    server = ThreadingHTTPServer((HOST, PORT), RequestHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        pass
    finally:
        server.server_close()


if __name__ == "__main__":
    run_server_silently()
